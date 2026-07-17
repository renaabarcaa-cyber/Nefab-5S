"""
Nefab Operations Hub
=====================
Sistema web unificado para control de operaciones:
  - VAS (Value Added Services)          [activo]
  - Inventario de Materiales            [proximamente]
  - Registro de Calidad                 [proximamente]

Un solo archivo app.py por preferencia de despliegue simple (Render).
Stack: Flask + Flask-Login + SQLite + openpyxl (export) + Chart.js (CDN)
"""

import os
import sqlite3
import io
import base64
from datetime import datetime, date
from functools import wraps

from flask import (
    Flask, request, redirect, url_for, render_template_string,
    session, flash, send_file, jsonify, g, Response
)
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# --------------------------------------------------------------------------
# Config
# --------------------------------------------------------------------------
APP_NAME = "Nefab Operaciones"
SECRET_KEY = os.environ.get("SECRET_KEY", "nefab-ops-dev-secret-change-me")

# En Render, montar un disco persistente en /var/data y setear DATA_DIR=/var/data
DATA_DIR = os.environ.get("DATA_DIR", os.path.join(os.path.dirname(__file__), "data"))
os.makedirs(DATA_DIR, exist_ok=True)
DB_PATH = os.path.join(DATA_DIR, "nefab_ops.db")

NEFAB_BLUE = "#144E8C"
NEFAB_ORANGE = "#FE8200"
NEFAB_GREEN = "#6CC24A"
NEFAB_GRAY = "#88888D"

AREAS_VAS = ["CARPINTERIA", "OUTBOUND", "PACKING", "WOODSHOP", "QUALITY", "OTRO"]
WC_BC_OPTIONS = ["WC", "BC"]
TAREAS_VAS = ["Stock Transfer", "ODR", "Pictures", "Quality", "Net weight",
              "Packaging", "Box", "Other"]

CATEGORIAS_MATERIAL = ["Madera", "Cartón", "Insumos", "Ferretería", "Otro"]
UNIDADES_MEDIDA = ["un", "kg", "m", "m2", "m3", "l", "plancha", "pallet"]
TIPOS_MOVIMIENTO = ["Entrada", "Salida"]

TIPOS_CALIDAD = ["Retorno", "Reclamo Metso", "Ticket", "Diferencia de Stock", "Reversa", "Scrap"]
ESTADOS_CALIDAD = ["Pendiente", "Esperando RDEL", "En proceso", "Aprobado", "Rechazado", "Cerrado"]

NEFAB_LOGO_B64 = (
    "iVBORw0KGgoAAAANSUhEUgAAAPAAAABiCAYAAABu+17aAAA5R0lEQVR42u19e3xcVbX/d+19zpmZPJvMpA9EAeUH2vouD6+vFAEp"
    "bSahXCY+QJTr57bCRVGatKWok+HV5lFqRcSCXq6Krw6vJmlBBSQ+uCAvURsUFUG40CaZpHnO45y91++PM3k1SZtmJmlasj6f+eTz"
    "SSbn7Mdaa3/X2mt/NwWC9ZUQxnehHQUQYSIhVhCWBSdV39FcfR1KwwZaIg4yldAOiWil8pdvuY6ksRZOKgGQzOCJGkKa0OqpDu9b"
    "znGfXXsasXwo/XcGQDiSQqwhTYMdfUGs6epHEGaBCGkgLICI9pffchw4+RQRcsGsDzovMyMMYsFMttT2GW27rnkR4bBAJKLHfDP9"
    "+wUVm0502HoC4HxAZ6UPxKQgpA+sHulorDovPY88VZ0LlG/5EqTRACeVYIKc4TF9lUDPE8TjGvZDscb1T41uX0gDdMi+GYAhSYh8"
    "Bh1cr5kAdgDD/HrJyvrW9l3Vd2fNiAGw1h5hGTmsUj5QBpPNzCQNYrZzsXgPAwArFmRQIWaLMDNJi9gZOAvAI3i0RgAYMgZWKSLJ"
    "8yAMH1gdcX8z6PaE6YG2+TQAL6J1yfiNehQCgFZsnSss33xODmiQFFlyfIP+d1ng/IZTOh6oemFCRzIpV68lmYbJSBkEMbODLMQp"
    "JIxTwFxBjrohUN7wDDP9TCj1o/Zo5V7XkEMS0ag66GNAmsEMMDSYecIPGNCswVqwaf5PyYq696Il4iC0Izuei6DBDIDUQdtxqI/7"
    "/25/Bh9NcsR3DtHPmfiAFLQGEd4DAJjfOo6nJT2peZmxNsOGEMxCLAYAtO0ZX+GXDY67DkErRroT2fkAYO2Q5bMg+ZMjHMYUdU6k"
    "2wadvTZO8qMdZjup2ElpkJBkWKcLy9vApvxDoHzrtQvOrc9FNKpQGjYObsCjllgc7AMQBLRiIsplw7h7fvDGBYhWpqFf1lanQ7Xj"
    "4B/3/8d3EQBNop8z8GHJ2gHA7/eX1+a7Xnaids+G9oLALKA1EfP7AQAt0OPOXSSii1dueROYPsDKJhCJLCMBAWUDRKE0AlQ4KoUI"
    "BAmCABhspzTbcYcgFpBp3aBy5OP+8tqz3EUyJCeCYWIK7xXspBSZ1ts0PD9BaIdAeEjR5mSyk6cdAGKRUMZb3dixhmZ/mxWYeYmL"
    "uiJ6jFItq5EAICSfR5Y3H1plH/8TBKsUQxqLi/JzTwfAaQWfQo/YnkIEPU3DCwHAYHaY7QGHhPFOEp6fB8obrkY0qhAOjxvjiinO"
    "pWQ77pAn56xA8p+3IBLRKK2Rc4Z5WEhDkWEJJdMwOhMoODMKRm4+DccF7H+9LZ2wGq1Qw6vyBW7ii6bJPEiRNIUQ8lMunF98DC0e"
    "RAAZ7CQ1tGOQlbPFH2zYhkhEIxQS2TFg90UGpwYcMvMuLwk2fBEtEedQeH1ORrt/kIRgfYYbBy/h2d5isGIhLS8xTgKA0YksJiCi"
    "5wdvXADNy9hJEcDT5dQFKxsEvcpfXpvvJlKPMQRI5ELr1IAjPLlf8gfrbkY0qg7MOWXq9SXbCcXS+EbJys3nZTWpdcyvwBDQDkBY"
    "CgCIVmo3FJnVWqUgDDDz6WMSWaGoAECKzbNd+KzVNG5/CShbk/QeD20uA0DHJgIkAiA51W8LT95X/GX1VyBaOcqIMzVgAjMRWLBh"
    "/SRwfsMp7gtCc0Z8yJFjSieyFhevvOFNAHjCrZnZA/sJzGCId7qoYUT23N2yYyKxykWB0x5dMgQxSF0MgMdNqh0jmgJmg1NxRdLc"
    "Gii7+f0jbUxk4fGClaNImkUwxT1uVnWHnktqHXrgwIohzAIi79vHrGizdUVgBQK/8+Tl2zyIRtOJLDf7vODc+vkAn8WOjWmEz8N6"
    "56SIgPNKQnUL3aTaMapzRATWICktkHMbSsMGoosZQJZS/ESSUwlFpvVOgvwhQGlINWfEB7dgUiQtgPi9AIBTj6NZ3mJi7YAZJ/Z6"
    "BhYBYITDNASfvXQ2mV4/tDMD1WNE0EqR5ZunkmIFABzTiVQiyXZCkZV7RklhwWeAiEZoh8he5lOQ5FTCFlZORaCsbhOilQpL18wl"
    "tSYBSQG4MeVfX5v9iSzNWhhmrlbyFABuImtxiN3ecEXabnmmmgNmEMPNRh8ujNbMR5sVQzvMzFUIhS1EK7XI8niaOhV3yPKtLw7W"
    "fQ5P327Pisy0kDxzSnX4kBSg95+8fJvn6ChKYIaQ0EK4BR0vdglESBeu3FQEUCkcGzNWu00s2EkxCfEhf3ntqUBEIxwWOFaFINix"
    "mQxrcVE872wAPA2d1ZKVrYW0bisp2/zh2ZCZlixSbmklzb4JUQ6IxPFdhn3S7HQyY3wOu7bDbiKrJCEAQArjLGF4FrJWOl2UMDMO"
    "EKzI8PgI8qIhRHCMwzYICQlUApiOgXYrdgjkZWn9tKhs81sQrVRZLbc83C5rdYiTGkd0PhQZpo9EOg6e9foDgtYAcCoAwpmdtqtI"
    "dAFIMGiGs8EMAa3AmivTp4z0EZzrCQq4sw6jwYQzEdrimx6jIhKskoqk+SYpzR0nfPZOb7p/c0mt8eacBJjpjKMDxrFgd//6lOKV"
    "W45DJKL95bX5DD4HKkXTsygcClamWEj5ruKBV5YCYIR2HKHFgsitcU7/dBECg1kBWYq3B1Eb6JRA3HnH9MWnJCQ7CYes3DP7Ozu+"
    "C0QuQSkMtMA5yu1Nu5MxtaOo4wyUglZgjFiBlU2QRjbbrIaPQXPGGgRWgJD5JPkUAK8B8ixhWovYTs0gfB7VJAXDY5COXwzg90fO"
    "E7Ma0g+GACBIGBKmJVmlAOXo7Bzu0AxpSK34xGlOMKXLLa2ciwNltX/paF5/QzbPEB8RH2uYwt19yxIvALOElCBFZy5YVT9/333V"
    "bewxslhFTCDDlFnNK7EGWTnggZ4PAPgVgVeBJNLwWRwB0xFQNgi8asEl9Rv33VXZj+x52EM6dDJMwSr1V6kQsiU0adIsWBi2ko6J"
    "HGE754Lk5WR4jmMnmbkRMzGRICK8bSYyxO5KbPquLy7b9Hxn8zX3HKVGzCAidlJ/AtCWZg3JjoI4DICkrRBwn50942Voh53k4wTY"
    "WQwNFSPuZWI+PrTFl0joj7MaLN44AlGSCys1GZ43q57kMgC7Edoh3NzLQdYxgSxtwxAAJPftqv7zBF94oqhs8w8E833C8LzfPQOc"
    "kaNjkIAgFBozMryaBUNpkp7vl6yo+0f77nV/mAzbwCwzXybDJM2pr8Yaqxqn5yVuoo8c0pCZ+gZmkCAwetjuWxl7MNIzHS0uLq87"
    "Rxqe49ixddbP/h5uf4UEO3wJgF2TDPM4i/ohxt3Cam0lFJ0jum5f86/CYP1FplbPQlABWHMG221u/YAGZmaPliCgHS0MK5dNRPMv"
    "uOkDvdGNMYTDAq1H1zKsWXgBdovnj4p9W6KUActdJsIERLKjtKVhiZaII0AXQhgA2UcGPo9KZiVBoOUlK+oWurQ0TJPhlcqaTEjt"
    "4zJrdDdV/zNQXv9jMnMvZ3vAQUb2x9DgHjGDeiRYpRwyrJMttn6K0rCB1iU0++t/D+wGM0CcPv6X7c80xe2e9PNrstROBloialFZ"
    "OIeZz3fh85E+z5wurTR98yDFctfJzLbSSiYCHnHJHDKINYjJZTrCP2Z40MlgO+4II+ecQGHeNxGtVCf7Oo+q+lVm7Rbvt5Wk6XkO"
    "93MMiFv7zDZyTheGdSK0zUfegIdniDHF0soMPftB/z5/CQPEmtHpnkLL6Cw+QTtKsHjpCJQ5ksH2gENWzuUlwbrn/9607hZXsRuy"
    "/J5p2uJmoQDio2k7jE1247PWKGFxePIrfaSGDwZBmUQ5CZOhHIXMwzEemRGaMspzUgSB0sAFDad03F/1wiBV73QnSMAwBulqMV4G"
    "vG0PuSvwlpPJ8IBTAxo0Bfrk4az337w5ovUI1Smz5FRCQVpbA8Haf3Y0UbOgLUbWBpMITOwgEuFsz5Mgtvzltfm2xzAl7MOKgfUA"
    "i+69xX14eo09kwtSXzzVPzXq1cj40aZ7qNxC4qUyKIfAEJknn4lctc9oygisFZk5Hp1KrAJQi1KIGVmJCQcJB5mAGgDE4LqL0uyr"
    "GSTrDGZt/++r0bXxI2TAlGaPZIK0fhBYdfMSOLoHWUtiMgDSWY0ryaVxYeJtgLzRSmqJw+ICZ2V4vEZgUc+aDuCBEd56GsdYgwgF"
    "fuT9koINKSbQZIaEQArS9LF2vh/buXb7qLamt2cCqVfPhDRPYW0zKFPzJTB0HBpEQngzM+LBUl59EUKhBkRr1ASOCMTZNGwmAAZK"
    "w0Df64S8RcOdaIFCS8QJBBs+RaZ5LtsJjSkTyRNBa9KKo8BMZaEnMAgoW5GVW8Sp+I+Y8DhphVlwC8FBHQORUUJEJYf/rxokLCiZ"
    "yJnZNgtDmNYHDwubsgZZ+eB45xPD8G/Md84nwwfW8czgM7MiyyfZTtxPhP1ker/AqQyeSSBWKYYw3lecPOO0TtATwzdfHPhdhwFP"
    "VpQZTDZaIolx/1wKIzDv5s8BdAsrh0FMU1qBGYoMS7AT/3OXr/9hgOnIHvUjITk1wBDyLLD+N3aSOCKleIc1iA6znsrWBCsoRxKE"
    "mmmnw3ZSHRYYITiw4wYT4uPAZ43V203s7Q5CO8gKfCYBofl3GvgLAZdn+ExynYIlKak+AeAJtEanb1FIkxyAeJE/WL+FiG1AxAfj"
    "VYAXMNGHSBrvZWUjo/1fAkNIYhKbEY2kUArjyJ/VJSKwYqJModMMwv+pQEYGgyDcLPaMo53DrZBiEMl0Pe+whMMucXtb3/tAxmJW"
    "NmfscIkEp+JKkfMo2b5XOBXfDyHmuYo+VTMmcs8lo/z40JZrX41WJjB9pZUE1gDJgDA9V4/Ox6UtXDtw68SZpmy8zIpMn6Ht/t/G"
    "uvt/lr5SRolZYxR8tLEjvAElfdaWtKogwyMAZIYmmBVJi5jVn7oWFb3Q+eBVPRpoIWnBvdQtk3yFo4VhvS1h62WYiRNKrJntuMOp"
    "AcWp+PDHjjtulRrE1MNDZpAk1vaAVnQ5WiIOWlsJ03OgP5NAYk5mM/RAtFK75AxUNgSfMw7RJUDYjdvdzLxkcV9WNIGgISSTEhcD"
    "GGTNnG79NUAkR30AIzOUwgwIRaZHaOVc0bWr+s9uUtEtQxZzejknk5IwEwAuSb30LoJ4Z1bgM4hYpUDAzwd/Yyv1qLYTAxAi08Mi"
    "kp0kMfTKBavq57vbaEcZ3Y57EZ4m02twciDc2VT9/QN3L+YMeE4mJ+4VqGBF5WR6swCfoUmagpXzotcjngRACIfF/t3VLxP4f0l6"
    "MoPRwFBppVZiOQBC6Wh9J0hnNhsvSZPI9ElO9X+9o7n6OvcAUOWoMZkz4DmZnLREFMJhwYQKaJU5fCZoSANE+Pmr0bVxhEJi+H4o"
    "asrmbiJDXQKAh68+dUUrVrPVeiEMZuBFduKf7Ghad30aNo+pbXhDGDCJwXtq52RKMh8CAC94Km8xCfEuVqksFG/ApeZRotmNH0PD"
    "tcuG2s123AFlCKOJJDsJBsmP+MtrT0VkNGsl8yytTXcP7AtotddUeBjA4F78mLF4QxiwI7U9fTflHftyQpurJ46BlWR4zTR8zkD5"
    "mSGkgJN6jXI8vwbAiIb04O0KHe8a+AeDnyDDzBRGu6dPDK8XEOVuKHAU6Hz6ClUyrA+mDPnn4rK6j090eeAbA0JrOZfhzkBOXLbM"
    "vWSaqQKsXGbKzDRUkbSYgd+0R6/sc+/5STvY0hqJSEQT8KBL05Op4yVAKxDzJxAKuWe4j4YjrETEdkIL0AJhWLsCwfrK8Sia52Lg"
    "OZlE+HuW44+f+f+I6DR2Uph6He/QqijAmgi004WHI+73HYxTFT3ATsqBW3CeAYyGZJViEub7ipNnnAaA0yHBURD7kWBla7CWJM0f"
    "F1U0nJft2wnn5A0iRHo5Sa/pUqRmmqCRgp1kFyn9K9dDjEgupU+QdeS85Q/MTitJk8AZV1ApmJYQoBAAnNxbPE0rMPPYT4ZtJxLQ"
    "ihksJYuf+strT0W0Ug3G8nMGPCeTyqqAUIFsHN5haJIWGPy79t3r9rqKOOqoI7u371UqAM0QJjImi2cIODZY64oTSu/0/j1NRk/Z"
    "zosIg0Z9SLoXCvAg3WwGRqxsRYY1Dyz+G6u3m25V3JE+zDAns9xuXQfvL9tyKiA+wCqFrBDXEYEZP0doh0QrJEI7Rq/qva9LlIYB"
    "xsPQqY1AhpB9sLTS9Jzcn99xFiKRB9zfZ5PUjhUrpxNE7B6VJSZAglBMVk7mvNBpnnXpyf1gYG/P5R2Na76J0A45Z8BTWEGmuCWl"
    "wJzm1JrpNh8m7CUoMA/WWREJnEemz8upATUlFonR8FmyHe8xYEbTq6wad6wAxIBH/OX1fxTSevdwPfGUjVhDGIIp9UkAD2RRHxSZ"
    "lmQ7+axQHESuD+gfAEmTFSWlYgoYqcTHQHQ1mZ43s50BLzRDsJPSDLr2uAtu+tFr0cpOY1oVPZPTF7M1FjQsgancyspakpkDPZCw"
    "ZrjFINNzeKeRWEuyfOCBAa8bw3EQnA34TExSklbq71onTympaHgvMxsMkoxBx6AlMXsFkQkIaNbdWVJIdxUELV8Y2layN3pVO0S2"
    "ToYRQCLRvnvt3nH++BqAPxav2na3UPa9ZHhOnzIvtEsqoYQnd34yoS8D0GBMl/WSYQnWCunrM4+NtZcEsUr+Ag79kwnG4cQ1RGDW"
    "2pKs/w5gJorr0+GXttlJPEKMJBMETYpmlRQnevIY9MTC0I0ldpI+CGVnflY7zWpCoPfCtH4NUPoGIRozWIM6Q8oGZ+PdIIJylLB8"
    "81PJeBmA/2HNEjJr2jF4of2IgosaQhhAFEbnfVe9WlixtdJk51kIKpz6uWAiKIdB9Gms3r5tGgyYmYRJ7KR+D6L/B2EUpV9IR7n5"
    "MhkGOSn1ja7mqswhWGTaidYGid17Y6l/VeDBW5JTeYq/vP4LwvT6OBXPED6PTvekM6tweaImGHGX30xkbQUgYjBYaHwKwJ0gmeU5"
    "IMZgnwb74LL5pLB0u9m9c81LgWDtz8jKXzNlXmi3yANE8j3Fr/ctnQYDJkWGx2DlNAPiGRJGM2utAS2OCTgtUJjeh5OYSkF/NKRn"
    "lGyciHJ9xYX9oR0xtJUQ5rdP4t1RoOgcgdvXONBYOU2M1YO0Mgcpysy6ukh2kgRJHy1cualIQsVnTCXzXmN3hd7SAlZrkEkxKrMi"
    "05BkJ8umBUIzazDw5ljj1deXBBu+RJbvm2wnHEDLo92ICUKlN9MxvaR0WfQ57HHbPNmbCsJhgcga219eexwxlWYt+zwbpg/sCDPX"
    "Y6L/31ljL81UeJfmhYao7WatAGIxdQc1+H/0/mmbFCJyEA6L9qaqW9juryHTZwCkMCczj/4d6/A0JV0vzCw/TpY3H1odM4kMMIhZ"
    "AZo/kU6azUzX3PJNgqZTSFrIyBYIbn4JOGV6t5EiEY1Q2OqIrosEyhvyyfStZTue4Z0wR3r+IVEaNtJ7ldmd+XYIhMIOnsrqkiPy"
    "fCb1Hc4/LYNGCyDAq7JQBTXrIBQ7KQB4Pwu8HU6Kp59IcZAXGsxEFxE7QKbnuViDwP7pN6Q2aIR2yI5oZVUgWJ9HVu6azC92OpIW"
    "rPanr0adnsPgEQDLtxGkfaQ6SIiQLllRt5AJy1ilaNYzhR6uCbMGhCgG8xVg5YDIzMpzQzsk2h4lzN/Bo1beFnLQAsdfVn+FMLwf"
    "cnmhM0oIkjtTZE7jPnD6rOX8Je5RsXBYdESqvxCo2JInrJyLdbLfAR1FlWAEYq0giS72BxvezWBJyCIxOEETDB/gPJpI9T3vgXVk"
    "jCYUFYhCsxRnkeEpYDuZzezzLHLEDCJxUramkAB7opzIorJwTkoUfomIbnT3gLOTByKGNqZR4cWoV9UwAIiO1/Mu8+/tLRJWzgpO"
    "DRxNRuwy/kvPJe4582wrlAaZudADsU/kBqynnZ4jtOq5+9MMgQvdTVocu+eoM6KuHXbs0Aoa+nh/sO7yQb0nhkkgnwadlCIqFYZ5"
    "CjtJpLeZshDJETG0PXPGQ8Tu/bRrbKtseyhl9z0kLN+/ueV54qjx8OykVPaVOk0b6nT22sL4hdzn5MILPfNpIxc+L1j17fnKGTiL"
    "HRtgEjh2uRCyMcKCtQMS8q1kWN8+0D4FCNA20rBZZOeVDPcaIvXKDHt5l9Lk9eY1A6n+ZFA7qT+SmSORTqkdJVMu0/F79j4MQdI0"
    "mPmP3Tu/sj+VQx4Q6Rnvm3ttKLSKl5Lp9UM72t3umJPJrOZpHujRvNCpAcVOBocYxrdfTUIyCC/M/OSkeYl6H94Y0yoV1Mp+gUyv"
    "zPyc6dG8DhCDJAj0MABo2FlDJMwsVX9yks+LAgAUuBwEPiJO5GhezQ/khB76ZDscSl8OqPHotBkw80E8dySiEdohu5o3/Iuc3jLW"
    "6hUyrDeyEUt2ElDCJTATyncEDIcJ0agqrNg6j8DnQtmUFeL2Ocn6REEIwU6qV7LzoJgGy6Uh+D/Cq4919pUKpWGjY9fX/0a2DrLW"
    "MUhDZkxiNlH2iZlnZSmCC4eItfqXNcDPAQCZRww+k6WdjwnpWcBK6WNs++gYMV/WZHgYwENtu655cfomaDJsB2mmvfZda59TdrIc"
    "THESksDZvZBZsKF41mZimCFNAPTYvl9W9wMA4vEjF8kRrYAbX83B59kbb5MS2D5ilTyCkjbirt0bHoNOVjIJhhDI5mVnrJ1ZXgZI"
    "IMIvhwG1yTPegGil8pfX5hPr5axswhzd0mz09Yosn9R24qGunVU/RzgsZscktUQclP7K6Gha38wq9VkiQSCh3wA3FrrxjB2PQ+C3"
    "0+QaDj2G6Zv7iIyPkOF5E5Sjs2bAzGpaPm88on4GCbBjp4QWVQCA1tbpOmHCAPjwMqktZzkoDRuxpnV3aSd1OUlTAqSP6YliMAmT"
    "wLyn4/4n/pHdy7docCaENia3FcTgCyAMZBM+k+mV0/EBGQS8oWzYEZZPQjkb2netfc69JymqZlcVVBpOx5rXbfcHa/OEmdvATtJJ"
    "O4Njj5zdvQKTQPQwEFVYut3E09mMPRkAC1ZGeuxqxrmegwlRUgvOrc91NJ8LZbt1WNmpN1DsJB5hIAnO2jFSImgF0PsgjTdBOzjW"
    "aJvGQTG28OSZOtl/Z0dz9db0DYUamI0HCgaNuGn9Fn+wYZ6wcr56VB9+OASAhlYE8G+OWBvc2mdl+/jfpPScmDF5nNsvTYYl2Ek+"
    "3dFY/fHpaLa/vOFiIcy7WDkqY6L5WWy5ACnhyTV1qv/ejp7e1UBYpI2XMWuNYsiIq74WCNYFyJP3BU4N2ADMY2h2NIQUrBIdphZP"
    "AADe+pDG00emMULThTANgOwsxL/MEBJMuBsIC4RgYHGWTm89CoFl0OJJ/TAj0Q0hC7NS0zwLE1YQQpLhNVRy4I5Y09o1acoeAiJD"
    "KGoaDTjDm99aahRCS2RHtPJyf7BhnvDkfJJT/Q5wjHBZMzMZJrSjfv96c1UHwITFNTMc1Lnw+eTl2zxdZH8c2kYWklecvhXQYUc1"
    "AxGNxWEnixxgGsvCon33ur3+YMNjwrDOZzuuj5ETU+zWQRCR5ZPs2H06FV8Xa6q6DVhL4zGqTB8jxyBN6OLFU1RKYkT3MJgp1tP7"
    "GZ2KN5OZa4Dh4FgR97DPQwDIvdTrCMBnAPut1Jkk5dugbM5YJxiaDA+Y+YnO3fG/AmGRdQK/NGMIEd139Ie/zMOZdddwSRqCneRO"
    "5Tj/Fmuqui19+RvGo0MywKThHk7PDgwhqPSzsmBoEY0aCLRElHFu/ScV4jvJ8p7twmkS4xaLpN9PIwjnSBjMxE5aOafez2zdJ0vE"
    "IALbKQ0tWzB4+XRL+s/SYkIiPangKb136IaASRDvMZWR9DislQNmM8P5dwASBESBiEZp2Bh191FWQiz3eRbxgykn0QshfWDtYLbe"
    "9ztRLo5AEAaRMCRIgO1knJX9oFDqtvamKrcuwE1YTTiHhgZ8hukz3KPBWeg/a4NMH9hJFGaln+nDD/si1f3Fy8MXCip4SPgKT2fH"
    "Hj/5yNog0wvtJIvSHWLNbAjDyoE0srArxVlxuiQt6GTvq14fPe/2s4bTWWJoSkoJWUCmD2AHmZCfkVZ5nIrnAADCwIhVnhCtVCWh"
    "b+XpZPwyIaQBw2NktKIxQEJIbSdAKd0MACMdU/YkooGweG1n1Sv+YP0fZE7RR9geSB+xO0pyU8yuz1GqDVr9mRk/ZymbOu//yvND"
    "4U24hhA5OHGiwRae0snerxLS12lkIUbRTsrDrJ8DANTUMCKRrBhxZyTSs3D5jSsVxH+AFY27KjKxtpMmE788aG22j9o8ieQGOLYA"
    "8dRWNAho0o67R5rZOBELDa0saH7u1WhVfDi2YQARGPu9XbowcSWnBnJAeorvI3av79LaNqz2YScRGZHgJKB/IA8G1alUd8aeiVho"
    "EJla672dD677x9DcTYe4zog000aR7P0IazvJR0ntNjG6tUYnBL1q2akX9z54bftwv8ICrUsIUVKIHHMb3Tx3UfecHIPCLp9W+PAL"
    "edzMVmlN9jN481sZ0aials5Opr3zl/CI2IFQGp59WcqDjVFpOHvZ9paIOgj2z/7YjB776ZVwWAwmtY4qmb+EsXgPu6iI3mhloXMy"
    "J3MyJ3MyJ3MyJ3MyJ3MyJ3MyJ3MyJ3MyJ3MyJ3MyJ3MyJ3MyJ3MyJ3MyJ3MyJ3MyJ3MyJ7NYCAgLtyZzKnXHk/ofyl5bs/IeOuw+"
    "hjMZo0m158BPpv9/uM8boQdTna9pH6fpHsexzzv8/mRjLobH8qDvDe0YWw8bDotxf5+xzJChh7PI8hjaIcdt97SN0ZF4T7qofkzf"
    "Q9I9VD7JcRpv3EMhmdX5yLSPk27LBGNyMJ3Iqs4daAdD7aFxjaVkRd3C4uWbjy8JhfMOdzXMP/sm/8GMb3EobB32JC7f5hn/eYda"
    "HQbfc6gBPqRi0shnFIS2FC+o2HpiUUXDmxecW597mE7pELqS6TOYgCkTvA2/e+lqs3jlDW/yl9ced7I7/pNzqiPm1l9em79wRf0J"
    "RWWb33J8aItv8k6V6TAcPE38t0k4nEO1ZdTfmYpXbnlTYcXWE+cHv7Fgcs8JyQznEgCwMLStpHjV5uMLK7bOm2gQ2F+x5QoB+jIz"
    "nwwwMXOcpPw1K+e2WNO6nYPfG9dLRCtVSbDhc8gp2qzjXdfHGqtuHcUkEHZpVfxl9aeT6b2b7fgtseZ1DeOzDbhnYxcuv7HE8eU/"
    "zHby0Vhz9ZcQCkksXsyIRPTCFfUnOKZ8mFk1xBqrvzP0nPRPf7D+Y8KTd4dK9n6js2ndLWPek+bUDVzQsBQwosTO5e07q34+TnuG"
    "+lxcvmWVJHEVgz+INLkea90vhLHbUYlvdjVv+O14nEWTUqJ0n2yJnSSoiJk9I4aDQNxDMJ7TpH/SuXPtPRM9wx+s30LSupBVyjPu"
    "dZbMmgzT4mQyFNu9/lfD/XXbXXTO5kKZY60H8X+COZDufjeT2K21c2tX07rfTagHg3q0ovYsMo2rWPPHhRA+BgCte0GiSTvqm527"
    "q58YbO94fQhUNERZc36sqXr5+EpNjOXbPH7LfoygH+toXPfFMfNfVreGTN96OCmDia3h5hID2EsQj5EzcEf77q/+Ydy2jGjPgnPr"
    "5zs+cRVBfxYk3uQ2Awyi58H0PVPv/87rzZGBobal21AcbDhXCHkbtGMxyANiHuVvmDVJ08vKvifWVP2f49lLYGX9J8k0q1mr97jU"
    "ypQCxFNEuKP9vft/gEhECwAoDtZ/S3oLbgX0C8zqvzTweRBth9ZvN3JL7g8EG34AgMfxNIRoSJ9QeqeXgXXEagGBqv3ltfnYERo+"
    "iB6JaIAp9nrBH+Ak4xBiw/GhLb5h5RnpEFyeJsc0Lxae3HdBiPvSfxj6iuPVljC9byON+QCAtj2jnsHAPBLirdLwfLNkRd3yock9"
    "QBxWOcLKOcnRonCClQAnfPZOb6Bi6w8NX+G9zM4CVqqelbpUa/VFIopqOCvNHP9vSirqbwCIpwoTlRQmSfM9YHQS4W4wGgm0k4B7"
    "ATwJqA8Y3oK7A8H6uyaCbgS8lcAnEnAvEd1N4HsJuGfoQ3QPIHdoyW0AgOieIXqhRWXhHCPP20Sm5xrWaCTGZcx8BTPdQ+AyM6fo"
    "t/5g/VoAfMBYUhpec6B8S63MnfcIM38AQt+hoC9jVpeD+B5A/7v05TweKKu7evB2ynEHQvPJBJx6sLE62TXFd4LpLaPmP/2ThFwg"
    "DOskgH9D4HvTY3k/Me4n5n0gvhzewmeLV27+rKubB8zZoAGt2PRRnWv9QZiejQCeBPOXWelLmfl6sO4X3rwttix6qiR484fGrMRE"
    "ecLKeRuAZwH8lJjuGzMXJH5GoCdHzMWIxa62WuTN+wlrFQd4rWb+HENtYahc8s27M/BM3kP4wBafUVxef4a0cv9LDXTfEGuq+tqo"
    "jpSGqwNkXENEr06A0QWipPoK6i4iab5D93euhuW7Xdj6IhDd6XqVNCdTaY1ES8TmNzVcL33z7komulcAuMf9/Qj+rGilS3COnqvU"
    "QOcfY43VvwIgEK1UgwNEmjRrWxPBnkAJWCcHNIi7YZp3zV+56Yy2aOWLB3pbglTs2JpAY/m7QpUC0ajq76z/nswNfNqOx67tbKy+"
    "6cCvFa7cdDWJxM0A/SkT8GubHtvQNjPzT2ON1bXjIZ3igZdvMHP9G/x9L/1vDHTrGMRAUFo5bbGm6isn99ZBzipyUtTwWWnlfET1"
    "91TEdlU1jvxWUWjz1WQnwsT6KTcW2sMjlJ0Qiajiss2bZa5/neqL3ZrX01v1ckskMeIR31m4/MZ1jjC+K/MCW/zBungsWnnbIBI6"
    "oFEDAFmTaHw/E+IThCIOO0lNSq9t371u74F/Ll6++XjhlU3C9N2+cEX9o3t3V788rBthgUiE/eVbT4VhPMBadUElTutorD6Q8Ddc"
    "Ul4XIjPvNpXqPxmg37lO1b2NUzDbYK2ZcUOsuerJSc1Fug0lK+oWsmFt5njPzzoa137ygC9u9JfVrQGJQjy+NmEQ48NQSgvQ97F6"
    "u4nkaxIvwUkfynY6gOuH33EA1Eh7DSKqZuU8F9u14Q5/ecMXWIgNCIV+gGho+PstNcqlc1GNOtHbyUpVAbhn+PdDcFwXHd/7MeEp"
    "PFENdF+adiQCLQe8myB4IgInCRAJwSr1BQjj28rwRbF0+wfQWpRGBZVpQ9cEAwKsabywwF9RHxTeok+rgfbazsZ1N405PD5/CXdH"
    "K7sAXDbhGE024lE2gYhAKEivThKDhHRtewjRSqcTuCZQcfNnIMSnANw6cgUdHha2Skq/ldeO9gT6jiPkvTYa7s5fwu68jIb6BJyv"
    "4/s7Y7vWNSIUtoAlQNsejflLuCta2Q3g6jF9DO2QiFSqwAXblgrTs171d3y/o6nqyg53zkYREux98Np2LF19UeD4d/yWDN+Wwoqt"
    "D3RHv/LSGAhLEMCkroIRmIhBkzWBSCiDihHa0Y6iLoGuIvcdRV2i8/Y1rwbKN10lPMUttrJXALgtPa86TdXD0PYdIBOkBpa1N238"
    "O5ZuN0eNZUvEaW9cFy1cuemh7l3XdLnjQkPIgt18hmCieSgNG/AVS8Q7x5IcLIMe6n9rq8uJZsh3CMNDSPT+N0I7JOJdHvS+lhp8"
    "b6x53fbBfzcI9BpJUyg7fjZuX7MdSK9qoR1yaBJGvmS0kutA+daPCsv7bifeFwJATPomw1N4d2DgjI93gB5wPXzEGYwPYtHKXn9Z"
    "3Xekr2BjoLzhfR2N9OzQSuJ6dja0vkone7vy/IFozG30YbE7EEOT6QFr+2/MqZD0BR7xH9d1ayy6ZrXbpxBPeG/xSDjDdI1O9nWY"
    "uv+6QZjo9iUsUAqBtj2E1dvNxV2vUSsARGvsjNkVmDSiIY3SRwkty9JjvocGx5G10wlQzpDXPjCRQ8Q5yHXQcqVzODG5Bl6RZu68"
    "wMpNSzui1zw9ap5Lw8b4hh9N24u9UTuUIKd3g5sljQpEK0ejmtJfGWg5y6bj6q4mT+5vzURyNYCNaF0ybVlc0il3LENRDC0m4RpC"
    "adggIf+PlaMBLhjV10ilKi6vP1N6Cj6iEvtv7Gja+HeEwhaia1LjoaK0A58oLwACK7REHIR2MB68aqwejyT8S1MwC1JdRESKxEcQ"
    "rfyFi0oGIfpomzSSMvULj93/ovDkfMcfrH8zE37c+b6+v4xiw1s2TlyXNjaw/XWVUi919vXdj3CYOltPvD+QeOUlJmwE8MAoVsI0"
    "9NI+a7tQqWvAvAbAF0YlDULfPEkrOh92YvPL378scShazYkTuhoa2t/ZuP4h/8raiJE/P+xfsfnp2O4N27H8ix5MSLfKBJAuXPnt"
    "IuKBM1ilfvB6c2QgPbl6yHAGqVJbgNbhpSkLWkcuxUoLOSN0gtECXVS2+S3C8C5mnfoRBle5lsgI2komMPTLLZ9LuqBgEsa7rEaj"
    "JQKp6XZm/Z9keh7xl9VfoyU1de2semX02I+KuwnRqEJp2EvMZ8OO/7Z9d2QvwhCIjONwW85yAFB747rH/OV1LwO0HMDGUSgt2wbM"
    "hnL5xUfMdcQN17isrkLkWILt+DNDyGRwaWc6m7XDWhg/S8Pi8XVlRAJw4ulMh4eT0eFIhAGmdqumNRCn38mcwq8Ggg3FIHy3o7v3"
    "T4hEHNfXDtuk0Xv/xphVXrsCSt4mPXnXsnauDTwjXuHy+scE5N3tba/sQiQSH9XQMAtESPsvbHg7iZyzKdn/RSyDXvQ0vK8v3ZPA"
    "M7nbpLdwq7+89rRYZP3TQ0aYZpfsinz5X4Gy2vtheC7Jv+Cma3ujlTGsXm0C0E4ieSVJSyuR+Pao1XBK+yJSIRwWscj6mkD5lvfI"
    "nILvFK3Y/Keu3RsemzgrXEOIgD2i781s5kjYA/8EQGjbQ0MJhvKbTyNBIVKOzcSCQQ5IstbY0dV09Z4JM5uY5FpYGjYW5R9nFfmK"
    "nB68IvfHOdcLPh2G8R2A+kC4EQAdyLfMmpJkWsWBYMMeoF4csDgrMjxeOInr2puqvz88J6QBpvZd9Fxg5U0XsJXbIH15t1Kq79ZA"
    "sOGvIHqEWd8da6p+JD3/lE5oEiIRXlCQt0hLo0Br/TzAhNYoHSTjTqCIRrloBfN7ly7dbj79NNmubtVkebuXYUhhLl293dzX1WcU"
    "4M0qtv9l07bEPBK4RHjzNnOip7nT+5ZH3Pkamejkt3BqgLV2XnUdAOuD4r1xfy0FKwcAfhgob+hnZjniqlcNaZis7SdjjdWfHrGC"
    "M8I1ApFISgS/8e862XMbWb4rAFzhL0AbBesf18K4D8n993ZGIj0AkwEwxRrprwA+5i+/+TQinM/Q5wK4gHz5n/AveMsLHNz82c4m"
    "enxIMdOTRLb+Kotke0dP33cQiejXgQE0Ax2l4W8FhLkRjA0ALhoFV12cT5qMOsPyrfLGnU/3Arfg9u3OorJFObYQn4dKNnY1fe2V"
    "qa6+IzAUDzoN1Wp/jhzj99KTc2/Jirr3tu9et1ce5DoO1pJAAAavUJ6/hPHoHjdOIvUOwLqUARtMIIIlvIULkNz/NwB7huKpwxBL"
    "a0NDJ0FU5S/M+4KNHtmW7CEGJbykHUAcB+BJHe+8KPZg+IVxnQRBENgB4c8aAmBFIJkOB1gBygdG20j4O6SEzNRBtBvLv/hwsXnC"
    "R4mwgog+RESXC0/B5YGKrb9gTl0Wa1z/2qisrXAYMECs0w5jx+RwBkE/fWB8fiijmKwLJKmlNOA4yV+//HoPQKAEXgZ81CvAORCy"
    "wEkN3GrG9Xo0VupBOt/R79eUn2NRdya6RwQw/sGEdgCSifRgkpWYTRD9bZxVWAOgtqYv7wNwYVFw8xIpPecB/HGASg1Pbrkm2uQv"
    "q18da6YmY2TSJtZ49VMAngJwfeHKTUU00HMhGeY3CNY9hRVbl3RHvtLtZg4rtb+89jgIs5KV88OSPO8SFdzig4aGgJAaca2dB0h6"
    "Ll4QrD9pX7R6OFkRjSogLDqb1j7uL294joivQGjHtxElnRINq6SVW6iTvVuRzcuqWmF0RTd0F51/3UWGr/gZtswfA/gYadJjXpOe"
    "x4SZfNVjs2bWJw3FNy5cRWxn9V0oDf8EywBEalTRyoYlSPY+K5wJsqKTmW+TNBwhCPwcmFpA8IERJ8Lpwle0zOnr+lJnU9UtQ1A2"
    "QnocqzDZsXs6mqorD/nCA7O/VEPurkJlshP4JdwPipdvPp5YrZbeeV/Tie4fojR8nht/uf+Wn8x5fb+R6mbwuwHwqAz1gYAoUsMI"
    "Q9AzeAeA14fyCSB2iZ4HncnBWSat/EWMxMsA08RhEDMAamRwF5gMCChicSkJsshOLO7Ydc2LE9WnMONl4clDIp44EeHwfrRGaRQM"
    "H7vlxGNiYM1MQkKTXhdrXP/UofDCuPUVi/dwV2TDHgB7ANxcEvpWnpPoPVcK4xaY5k/9ZQ1LjaEYYeS+3OI93B25pgvA9/xltSmZ"
    "U/QDI9HzUQCN6P2wAUST0PRfZPlM6L5PszT/Q9BwTpCJQUCSTK/UKnUVgC+jdcnwpISWEKIAM2+V3vz/8Sdf+WgM+BUxNuhE7586"
    "mqp/40K0bFGTLlEI7ZBd0co/BcpuukTkL/qpv6z+OjZwF42XzgdT7/0U8wQbniVplh0f2uJ7FdGUq2wRPZTManEtnq0tkkAGxNTp"
    "TZOKtCmkyZy6d9Q2Ujgs/M+Jp6U355qFK+ob957e9wpQg/FJv5lBJA6ahW6BTvdhbL+jGF1SGw3pzgfpVQBfDwTri8nw/FdxXsGC"
    "zsja/3MJyHfIv0crk/7y+l8KI+eCBcH6k/ZFqv85IjYfltXbDdxOdsnTmz9EOQUn6ETPHSN0QQ/qB7PuJIh3ILTDwuI9zhjjCIdF"
    "T+sreRBGLtjuGDc9zZrACoY3uX5vdJg0vSRYdw9k3q+1UtcDuDidWHMOTMoB/DCEJIb6DCKRZ9ys/BgDdh3eIXSUmYsmzEJPTL/L"
    "Q7ULI+ajPVrZB+C+krLN7eQt+A2U8wmjePm2Aq+333wtWhkbyiC2LiF8NuzFwBJbpP71D9YO02Dm88FOu3j5tgKS+kpO9jUJ6K9q"
    "ElJopYYhjJCCWelk3w2Q1urjLrjpevf5aQNwExfk4YKonerfSuD/mHdh/YtC5r0Tyb5LADAehTxcGHrwFadSoTRsdDRv/FmgvOF0"
    "6c39mk72m8wJCNJyTDFJFIqIaslbsCM+0HkNGqNfdzmpR9zo6CuWePCqJJJ6P1uDV0VkFLgBjIKhCc/vZEQiKVpZWwmPuccx5E/R"
    "uuTDo6qSDrccsxQCLTySi5gAcPGqbcd3dnZ2IFqZGFKaUFQgHvagOTLAxM8TAR6y5Vi0qDeTMEKKxM0AVg1eDzuUGGrbQ7h9jY2l"
    "q00tPdsoNdArOPldAMCOkAaNKMZgNIncopUlAy+WtUci96a3tNK6VSMRiaQS5bUXCG+x4aR6mw5MQI20r2RCzkdpuAvzIdBbTO1N"
    "V/3OX1a7wShYWBco2/RUR/NZW0c5mzQ6jDVXPeUP1rcIy3dVoKz2hx3R9c+McmxtewgtEcet+rrx7cY865V9d1X3H/Z0t+2h8cLE"
    "Ez4b9na3Gwv2RytfHkpWzl/CWL7Ng3inEqZ8QTm2BpBvCMveaVPOKf6y2ktjzfTwCKigAEBXbP2i0A6ZWj7mTnZEk1H3GWHmF9h2"
    "97UdOzdMWMDgL7/5OmnlBJPx1OcB1A2uvINbSq9HKwcCwfrvgcR/Ggo5Sg10GHG+303QjNgfnsC30UFjpXEuQRq8snTxnnWBZ+md"
    "wpO/ge0BVnSAo0gXjbRHrr47UL7lHunzf60kuCXe3lRTe8DK4iwqC+fYQq4R3gKp+/dnWsvMgBjedohepRDaITuilX/zr6y9zMgv"
    "uSsw8FJtR+O6KjeUGRfW2e0tV/YdIt05YpeDUXx+TYFw7CcChQX/Jyo2f74tWjlyTgcKK7bOI6Ivs5185fXeeJsLhyPsOtiw6Ghe"
    "/4w/WHujkTv/2kDF1rukcK7ed19128g3zltRf4Jhye3Cynu/7t//qbZd1+5DKCRBaRjszjc5rHYg0X0dmbnfLqmoe7E9uu4PIxNx"
    "/rKbTicz9xs60fPHhZ43P9rpJqAOyAUIBjMLIgctEceFuVcNbmPWByq2fFh4i24uKrvp8a7mjf87uowRQIRAsmE1mB+H6f2Ff0Vt"
    "KBat/NWYgpBgw+eE5bvT6e2/DkAYoR1yqBqMiAFmIY3OtM4c+rK/dHK4vyv/BsPjuaK4vP7Szsbqu0fonAIAVX7zOvJ4BCd1o8HA"
    "Vmh9o/TkPxQob3gSGo8AvA+CCphxgbR871XJ7nWvN6//F8JhseCxvBwl5ddVsq+lq2nDn1AaNrBsnJXyUYhY49VP+YN1j5FhXVNY"
    "sfX2brcggEbGSsKUtymlrzC8RReqeKxm3y/X9bsDSuogdYdElkkajjFBBCRImDQW0hIjGmaghuWqmku1jV+TJ/9UJLonSOnXwNS9"
    "l9oJSghv3k3+cr6MUL8TTH8jsMVE77JBq8gwSzjetc0W4hcAE1ro8KG/lkSmIK3i5njIIbZr/Y8CwboPilz/Wn9Z/XOxaPUPx3pv"
    "Ikgj4A/W354eZjogRlaQhods53vtzVW/RSgkQNALQ0jsS9F1RGITk++PJcGG3Uz4PYD9AB1HrD8D6SlWOlWGlkhidAVVhBFmEYvQ"
    "V/3BBiUs39cdR1/gD9b/hIieZ7Ak0LtB4hMgQznx/Z/r3FX903TsqEbNTTgsuiPXdAXKN4XA8n6Q9Wwg2PBjJvyJGYII7xHSqmSt"
    "X1Jsf7o1WplKV+fxAX7QIGkS7BSNhqV7GGBK9m36D0+B+Yw0cu8rrNj0ge5o6OUhRJNOenZEql7wL7/pPPL5fiJ9+Y8Eyusfhsav"
    "NXGPIFoEFmXCm7dY2wNNGuK7SJcVI+TqnCZtCghihUigvOH/GFqChxccItIQ0mStXog1Vm128xruzZms+WdEzpmGlRcNBBtaGfg5"
    "EV4FwwvgHOHNO0vFe7bGmqp+bcQaqxpPKA3/YqCILgPE50GqCtKQ0Aok5dOc7KmMNa6PojRsIBJxdEXDh4Xh7VXJgesBEOYv4TEe"
    "0A3CyW0n3QTD8w0j3r8MwP1Dk58uW2u79+oX/cH6Hyl74HxF4r8B0EESIQCAFLQytfM8gH3jQSiSRjc7qb9Km+PjxnqhJXJfNNI2"
    "P7j5c+w435Us+sbHs0SvN2MAwCUl5XU7QfIqMK4mwxTMDChnP4Mf1Dq1tXPn2t+7//aVKa29QqQcVuJ5MZQlPgA5hCHMpwurnWTv"
    "SSRoffHKLY90RitfG1l/y0wvCSH+BoXl6Xq0Axd4BZKWlqlfpicJAHFrFCkA2xecG75P5867UgvxSWK9AkICStlM9CtO9Xy1s3nj"
    "ky7wGeWgGBE3ix0jChcF634hSaxnos+QND3uOuR0MeEHwunb3Nm08e8TbrMNGc81v56/ctNp2vR9DYIuJJKfJje86GClbjFSbdd3"
    "PLipfczdw2k9YPA+Vs5fBB1YIuvOfW90Y0yu2PQZM6fwDiNpfBmgL7ttSjuCSEQjFJKx6MYni85Zv5RzF36FSF4GQ5wticDKAaR4"
    "UqV6Lo3trP7h6GSYm4WXJAa0k/wLgZeA6X1j7l1jUiBhQTslB2TfKdZc/SSAj/rL6i8hKdYA+gqQ4QFrAOIvOt6zJtZUdTvA9P8B"
    "d6lNaz9zPWIAAAAASUVORK5CYII="
)


app = Flask(__name__)
app.secret_key = SECRET_KEY


# --------------------------------------------------------------------------
# DB helpers
# --------------------------------------------------------------------------
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'user',
            activo INTEGER NOT NULL DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now'))
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS vas_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            delivery TEXT,
            area TEXT NOT NULL,
            fecha TEXT NOT NULL,
            nombre TEXT NOT NULL,
            wc_bc TEXT NOT NULL,
            horas REAL NOT NULL DEFAULT 0,
            tarea TEXT NOT NULL,
            observaciones TEXT,
            created_by TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now'))
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS materiales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT UNIQUE NOT NULL,
            nombre TEXT NOT NULL,
            categoria TEXT NOT NULL,
            unidad TEXT NOT NULL,
            stock_minimo REAL NOT NULL DEFAULT 0,
            activo INTEGER NOT NULL DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now'))
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS movimientos_inventario (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            material_id INTEGER NOT NULL,
            tipo TEXT NOT NULL,
            cantidad REAL NOT NULL,
            lote_ot TEXT,
            delivery TEXT,
            fecha TEXT NOT NULL,
            responsable TEXT,
            observaciones TEXT,
            created_by TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (material_id) REFERENCES materiales(id)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS calidad_registros (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tipo TEXT NOT NULL,
            fecha TEXT NOT NULL,
            guia TEXT,
            delivery TEXT,
            cliente TEXT,
            item TEXT,
            descripcion_material TEXT,
            cantidad REAL,
            ubicacion TEXT,
            cantidad_sistema REAL,
            cantidad_fisica REAL,
            peso_kg REAL,
            proceso TEXT,
            rdel TEXT,
            motivo TEXT,
            causa_raiz TEXT,
            acciones_correctivas TEXT,
            responsable TEXT,
            estado TEXT NOT NULL DEFAULT 'Pendiente',
            observaciones TEXT,
            created_by TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now'))
        )
    """)
    conn.commit()

    # Migraciones seguras para bases de datos ya existentes (creadas antes de
    # agregar estas columnas). Si la columna ya existe, sqlite lanza
    # OperationalError y lo ignoramos.
    migrations = [
        "ALTER TABLE users ADD COLUMN activo INTEGER NOT NULL DEFAULT 1",
        "ALTER TABLE movimientos_inventario ADD COLUMN delivery TEXT",
        "ALTER TABLE calidad_registros ADD COLUMN peso_kg REAL",
        "ALTER TABLE calidad_registros ADD COLUMN proceso TEXT",
        "ALTER TABLE calidad_registros ADD COLUMN rdel TEXT",
    ]
    for stmt in migrations:
        try:
            conn.execute(stmt)
            conn.commit()
        except sqlite3.OperationalError:
            pass

    # Seed admin user si no existe ninguno
    cur = conn.execute("SELECT COUNT(*) AS c FROM users")
    if cur.fetchone()[0] == 0:
        conn.execute(
            "INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)",
            ("admin", generate_password_hash("nefab2026"), "admin")
        )
        conn.commit()
    conn.close()


# --------------------------------------------------------------------------
# Auth helpers
# --------------------------------------------------------------------------
def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login", next=request.path))
        return f(*args, **kwargs)
    return wrapper


def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login", next=request.path))
        if session.get("role") != "admin":
            flash("No tienes permisos de administrador para esta acción.", "error")
            return redirect(url_for("hub"))
        return f(*args, **kwargs)
    return wrapper


def current_user():
    if "user_id" not in session:
        return None
    return {"id": session["user_id"], "username": session.get("username"),
            "role": session.get("role")}


# --------------------------------------------------------------------------
# Templates (embebidos para mantener un solo archivo)
# --------------------------------------------------------------------------
BASE_HEAD = """
<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{{ title }} - Nefab Operaciones</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4"></script>
<style>
:root{
  --blue:#144E8C; --dark:#0D3A6E; --orange:#FE8200; --green:#8CC24A;
  --gray:#88888D; --red:#E34948; --bg:#F4F6FB; --card:#FFFFFF;
  --border:#E2E8F0; --text:#1E293B; --muted:#64748B;
}
*{box-sizing:border-box;}
body{margin:0;font-family:'Segoe UI',Arial,sans-serif;background:var(--bg);color:var(--text);}
a{text-decoration:none;color:inherit;}

.topbar{
  height:56px;background:var(--dark);border-bottom:none;
  display:flex;align-items:center;justify-content:space-between;padding:0 16px;
  position:sticky;top:0;z-index:10;
}
.topbar-left{display:flex;align-items:center;gap:12px;overflow:hidden;}
.brand-logo{height:26px;width:auto;display:block;filter:brightness(0) invert(1);}
.app-name{font-weight:600;font-size:14px;color:#fff;border-left:1px solid rgba(255,255,255,0.25);padding-left:10px;white-space:nowrap;}
.app-name small{display:block;font-size:11px;color:rgba(255,255,255,0.65);font-weight:400;}
.topbar-right{display:flex;align-items:center;gap:10px;flex-shrink:0;}
.avatar{width:30px;height:30px;border-radius:50%;background:var(--orange);color:#fff;
  display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:700;flex-shrink:0;}
.user-email{font-size:12px;color:rgba(255,255,255,0.85);white-space:nowrap;}
.role-badge{background:var(--orange);color:#fff;font-size:9px;font-weight:700;padding:1px 6px;border-radius:4px;}

.auth-page{min-height:100vh;display:flex;align-items:center;justify-content:center;
  background:var(--bg);padding:20px;}
.auth-card{background:var(--card);border:1px solid var(--border);border-radius:12px;
  padding:28px;max-width:380px;width:100%;}
.auth-card h2{margin:4px 0 4px;font-size:19px;text-align:center;}
.auth-card label{display:block;font-size:12px;color:var(--muted);margin:14px 0 4px;font-weight:600;}
.auth-card input{width:100%;border:1px solid var(--border);border-radius:7px;padding:10px 12px;
  font-size:14px;font-family:inherit;}

.layout{display:flex;min-height:calc(100vh - 56px);}
.sidebar{width:230px;background:#fff;border-right:1px solid var(--border);padding:14px 0;
  flex-shrink:0;display:flex;flex-direction:column;}
.sidebar-label{font-size:10px;font-weight:700;color:var(--muted);text-transform:uppercase;
  padding:14px 18px 6px;}
.sidebar-btn{display:flex;align-items:center;gap:11px;padding:11px 18px;margin:2px 10px;
  font-size:13.5px;color:var(--text);border-radius:9px;font-weight:500;}
.sidebar-btn svg{flex-shrink:0;color:var(--muted);}
.sidebar-btn:hover{background:#F0F4F9;}
.sidebar-btn.active{background:var(--dark);color:#fff;font-weight:700;}
.sidebar-btn.active svg{color:#fff;}

.content{flex:1;padding:20px;min-width:0;overflow-x:hidden;display:flex;flex-direction:column;min-height:calc(100vh - 56px);}
.app-footer{margin-top:auto;padding-top:24px;text-align:center;font-size:11px;color:var(--muted);}

.page-title{margin:0 0 16px;font-size:20px;font-weight:700;color:var(--blue);}
.muted{color:var(--muted);font-size:12px;}

.kpi-row{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:10px;margin-bottom:20px;}
.kpi-card{background:var(--card);border:1px solid var(--border);border-radius:10px;
  padding:14px;text-align:center;}
.kpi-value{font-size:22px;font-weight:700;color:var(--blue);}
.kpi-value.alert{color:var(--red);}
.kpi-label{font-size:11px;color:var(--muted);margin-top:2px;}

.filters-bar{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px;align-items:flex-end;}
.filters-bar > div{min-width:150px;}
.filters-bar input, .filters-bar select{border:1px solid var(--border);border-radius:7px;padding:8px 10px;
  font-size:13px;font-family:inherit;background:#fff;width:100%;}
.filters-bar label{display:block;font-size:11px;font-weight:700;color:var(--muted);margin-bottom:4px;text-transform:uppercase;}

.card{background:var(--card);border:1px solid var(--border);border-radius:12px;padding:18px;margin-bottom:18px;}

.btn-primary{background:var(--blue);color:#fff;border:none;border-radius:7px;
  padding:9px 16px;font-size:13px;font-weight:700;cursor:pointer;display:inline-block;}
.btn-primary:hover{background:var(--dark);}
.btn-secondary{background:#fff;color:var(--muted);border:1px solid var(--border);border-radius:7px;
  padding:9px 16px;font-size:13px;cursor:pointer;display:inline-block;}
.btn-danger{background:#fff;color:var(--red);border:1px solid var(--red);border-radius:7px;
  padding:9px 14px;font-size:13px;cursor:pointer;}
.btn-mini{font-size:11px;padding:5px 9px;border-radius:5px;background:var(--blue);color:#fff;
  border:none;cursor:pointer;display:inline-block;}
.btn-mini-danger{background:var(--red);}

.table-wrap{overflow-x:auto;background:var(--card);border:1px solid var(--border);border-radius:10px;}
.data-table{width:100%;border-collapse:collapse;font-size:12.5px;min-width:640px;}
.data-table th{background:var(--blue);color:#fff;text-align:left;padding:9px 10px;
  font-size:11px;white-space:nowrap;}
.data-table td{padding:9px 10px;border-bottom:1px solid var(--border);}
.data-table tr:hover td{background:#F8FAFC;}
.actions-cell{white-space:nowrap;}
.actions-cell .btn-mini{margin-right:4px;}
.row-alert td{background:#FCEBEB;}

.pill{display:inline-block;padding:2px 8px;border-radius:10px;font-size:10px;font-weight:700;}
.pill-blue{background:#E6F1FB;color:var(--blue);}
.pill-orange{background:#FFF3E0;color:#b5610a;}
.pill-green{background:#EAF3DE;color:#4c7a17;}
.pill-red{background:#FCEBEB;color:var(--red);}

.form-card{background:var(--card);border:1px solid var(--border);border-radius:12px;
  padding:20px;max-width:700px;}
.form-card label{display:block;font-size:12px;color:var(--muted);margin:12px 0 4px;font-weight:600;}
.form-card input, .form-card select, .form-card textarea{
  width:100%;border:1px solid var(--border);border-radius:7px;padding:9px 10px;font-size:13px;
  font-family:inherit;background:#fff;color:var(--text);
}
.form-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:12px;}
.form-actions{display:flex;gap:10px;margin-top:20px;}

.flash{padding:10px 14px;border-radius:8px;margin-bottom:14px;font-size:13px;}
.flash-error{background:#FCEBEB;color:var(--red);border:1px solid #F5C2BC;}
.flash-success{background:#EAF3DE;color:#4c7a17;border:1px solid #bfe8cc;}

.hub-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:16px;}
.module-card{
  background:var(--card);border:1px solid var(--border);border-radius:12px;padding:22px;
  text-align:left;transition:transform .15s, box-shadow .15s;
}
.module-card:hover{transform:translateY(-2px);box-shadow:0 6px 18px rgba(20,78,140,.12);}
.module-card.disabled{opacity:.55;cursor:not-allowed;}
.module-card h3{margin:10px 0 4px;font-size:15px;color:var(--blue);}
.module-card p{margin:0;font-size:12.5px;color:var(--muted);}
.module-card svg{color:var(--blue);}
</style>
</head>
<body>
"""

def _icon(name):
    icons = {
        "home": '<svg viewBox="0 0 24 24" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M3 11l9-8 9 8"/><path d="M5 10v10h14V10"/></svg>',
        "vas": '<svg viewBox="0 0 24 24" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 8l-9-5-9 5 9 5 9-5z"/><path d="M3 8v8l9 5 9-5V8"/><path d="M12 13v8"/></svg>',
        "inv": '<svg viewBox="0 0 24 24" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect x="3" y="7" width="18" height="14" rx="1"/><path d="M8 7V4h8v3"/><path d="M3 11h18"/></svg>',
        "quality": '<svg viewBox="0 0 24 24" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M9 12l2 2 4-4"/><circle cx="12" cy="12" r="9"/></svg>',
        "users": '<svg viewBox="0 0 24 24" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M17 21v-2a4 4 0 0 0-4-4H5a4 4 0 0 0-4 4v2"/><circle cx="9" cy="7" r="4"/><path d="M23 21v-2a4 4 0 0 0-3-3.87"/><path d="M16 3.13a4 4 0 0 1 0 7.75"/></svg>',
        "scan": '<svg viewBox="0 0 24 24" width="16" height="16" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M3 7V5a2 2 0 0 1 2-2h2"/><path d="M17 3h2a2 2 0 0 1 2 2v2"/><path d="M21 17v2a2 2 0 0 1-2 2h-2"/><path d="M7 21H5a2 2 0 0 1-2-2v-2"/><line x1="7" y1="12" x2="17" y2="12"/></svg>',
    }
    return icons.get(name, "")

NAV = """
<div class="topbar">
  <div class="topbar-left">
    <img src="/logo.png" alt="Nefab" class="brand-logo">
    <span class="app-name">Nefab Operaciones<small>Control de operaciones</small></span>
  </div>
  <div class="topbar-right">
    {% if session.get('user_id') %}
    <span class="user-email">{{ session.get('username') }}{% if session.get('role') == 'admin' %} <span class="role-badge">ADMIN</span>{% endif %}</span>
    <a href="{{ url_for('logout') }}" class="btn-secondary" style="padding:6px 12px;">Salir</a>
    {% endif %}
  </div>
</div>
<div class="layout">
  <nav class="sidebar">
    <a href="{{ url_for('hub') }}" class="sidebar-btn {{ 'active' if active=='hub' else '' }}">""" + _icon("home") + """ Inicio</a>
    {% if session.get('user_id') %}
    <div class="sidebar-label">Módulos</div>
    <a href="{{ url_for('vas_list') }}" class="sidebar-btn {{ 'active' if active=='vas' else '' }}">""" + _icon("vas") + """ VAS</a>
    <a href="{{ url_for('inv_dashboard') }}" class="sidebar-btn {{ 'active' if active=='inv' else '' }}">""" + _icon("inv") + """ Inventario</a>
    <a href="{{ url_for('calidad_list') }}" class="sidebar-btn {{ 'active' if active=='calidad' else '' }}">""" + _icon("quality") + """ Calidad</a>
    {% if session.get('role') == 'admin' %}
    <div class="sidebar-label">Administración</div>
    <a href="{{ url_for('admin_users') }}" class="sidebar-btn {{ 'active' if active=='users' else '' }}">""" + _icon("users") + """ Usuarios</a>
    {% endif %}
    {% endif %}
  </nav>
  <main class="content">
"""

FOOTER = """
    <footer class="app-footer">&copy; 2026 Nefab Group &middot; Operations Hub</footer>
  </main>
</div>
</body></html>
"""

LOGIN_TEMPLATE = BASE_HEAD + """
<div class="auth-page">
  <div class="auth-card">
    <div style="text-align:center;margin-bottom:12px;"><img src="/logo.png" style="height:34px;"></div>
    <h2>Nefab Operaciones</h2>
    {% with messages = get_flashed_messages(with_categories=true) %}
      {% for cat, msg in messages %}
        <div class="flash flash-{{ 'error' if cat=='error' else 'success' }}">{{ msg }}</div>
      {% endfor %}
    {% endwith %}
    <form method="post">
      <label>Usuario</label>
      <input type="text" name="username" required autofocus>
      <label>Contraseña</label>
      <input type="password" name="password" required>
      <div style="margin-top:18px;">
        <button class="btn-primary" style="width:100%;" type="submit">Ingresar</button>
      </div>
    </form>
  </div>
</div>
</body></html>
"""

HUB_TEMPLATE = BASE_HEAD + NAV + """
  <h1 class="page-title">Bienvenido, {{ session.get('username') }}</h1>
  <p class="muted" style="margin-bottom:18px;">Selecciona un módulo para comenzar.</p>
  <div class="hub-grid">
    <a class="module-card" href="{{ url_for('vas_list') }}">
      """ + _icon("vas") + """
      <h3>VAS</h3>
      <p>Control de Value Added Services</p>
    </a>
    <a class="module-card" href="{{ url_for('inv_dashboard') }}">
      """ + _icon("inv") + """
      <h3>Inventario de Materiales</h3>
      <p>Kardex, stock y trazabilidad</p>
    </a>
    <a class="module-card" href="{{ url_for('calidad_list') }}">
      """ + _icon("quality") + """
      <h3>Registro de Calidad</h3>
      <p>Retornos, reclamos, tickets y diferencias de stock</p>
    </a>
  </div>
""" + FOOTER

VAS_LIST_TEMPLATE = BASE_HEAD + NAV + """
  <h1 class="page-title">Control de VAS</h1>
  {% with messages = get_flashed_messages(with_categories=true) %}
    {% for cat, msg in messages %}
      <div class="flash flash-{{ 'error' if cat=='error' else 'success' }}">{{ msg }}</div>
    {% endfor %}
  {% endwith %}

  <div class="kpi-row">
    <div class="kpi-card"><div class="kpi-value">{{ kpi.total_horas }}</div><div class="kpi-label">Horas totales (filtro)</div></div>
    <div class="kpi-card"><div class="kpi-value">{{ kpi.total_registros }}</div><div class="kpi-label">Registros</div></div>
    <div class="kpi-card"><div class="kpi-value">{{ kpi.top_area }}</div><div class="kpi-label">Área con más horas</div></div>
    <div class="kpi-card"><div class="kpi-value">{{ kpi.top_tarea }}</div><div class="kpi-label">Tarea más frecuente</div></div>
  </div>

  <div class="card">
    <canvas id="chartArea" height="90"></canvas>
  </div>

  <div class="card">
    <form method="get" class="filters-bar">
      <div>
        <label>Mes</label>
        <input type="month" name="mes" value="{{ filtros.mes or '' }}">
      </div>
      <div>
        <label>Área</label>
        <select name="area">
          <option value="">Todas</option>
          {% for a in areas %}
          <option value="{{ a }}" {{ 'selected' if filtros.area==a else '' }}>{{ a }}</option>
          {% endfor %}
        </select>
      </div>
      <div>
        <label>Tarea</label>
        <select name="tarea">
          <option value="">Todas</option>
          {% for t in tareas %}
          <option value="{{ t }}" {{ 'selected' if filtros.tarea==t else '' }}>{{ t }}</option>
          {% endfor %}
        </select>
      </div>
      <div><button class="btn-secondary" type="submit">Filtrar</button></div>
      <div><a class="btn-secondary" href="{{ url_for('vas_export', **filtros) }}">Exportar Excel</a></div>
      <div><a class="btn-primary" href="{{ url_for('vas_new') }}">Nuevo registro</a></div>
    </form>
  </div>

  <div class="table-wrap">
    <table class="data-table">
      <thead>
        <tr>
          <th>Fecha</th><th>Delivery</th><th>Área</th><th>Nombre</th><th>WC/BC</th>
          <th>Horas</th><th>Tarea</th><th>Observaciones</th><th></th>
        </tr>
      </thead>
      <tbody>
        {% for r in registros %}
        <tr>
          <td>{{ r.fecha }}</td>
          <td>{{ r.delivery or '-' }}</td>
          <td>{{ r.area }}</td>
          <td>{{ r.nombre }}</td>
          <td>{{ r.wc_bc }}</td>
          <td>{{ '%.2f'|format(r.horas) }}</td>
          <td>{{ r.tarea }}</td>
          <td>{{ r.observaciones or '' }}</td>
          <td class="actions-cell">
            <a class="btn-mini" href="{{ url_for('vas_edit', record_id=r.id) }}">Editar</a>
            <form method="post" action="{{ url_for('vas_delete', record_id=r.id) }}" style="display:inline;"
                  onsubmit="return confirm('¿Eliminar este registro?');">
              <button class="btn-mini btn-mini-danger" type="submit">Eliminar</button>
            </form>
          </td>
        </tr>
        {% else %}
        <tr><td colspan="9" style="text-align:center;color:var(--muted);">Sin registros para este filtro.</td></tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
<script>
  const ctx = document.getElementById('chartArea');
  new Chart(ctx, {
    type: 'bar',
    data: {
      labels: {{ chart_labels|tojson }},
      datasets: [{
        label: 'Horas por Área',
        data: {{ chart_data|tojson }},
        backgroundColor: '#144E8C'
      }]
    },
    options: { responsive:true, plugins:{legend:{display:false}} }
  });
</script>
""" + FOOTER

VAS_FORM_TEMPLATE = BASE_HEAD + NAV + """
  <h1 class="page-title">{{ 'Editar registro' if record else 'Nuevo registro VAS' }}</h1>
  {% with messages = get_flashed_messages(with_categories=true) %}
    {% for cat, msg in messages %}
      <div class="flash flash-{{ 'error' if cat=='error' else 'success' }}">{{ msg }}</div>
    {% endfor %}
  {% endwith %}
  <div class="form-card">
    <form method="post">
      <div class="form-grid">
        <div>
          <label>Fecha *</label>
          <input type="date" name="fecha" value="{{ record.fecha if record else today }}" required>
        </div>
        <div>
          <label>Delivery</label>
          <input type="text" name="delivery" value="{{ record.delivery if record else '' }}">
        </div>
        <div>
          <label>Área *</label>
          <select name="area" required>
            {% for a in areas %}
            <option value="{{ a }}" {{ 'selected' if record and record.area==a else '' }}>{{ a }}</option>
            {% endfor %}
          </select>
        </div>
        <div>
          <label>Nombre *</label>
          <input type="text" name="nombre" value="{{ record.nombre if record else '' }}" required>
        </div>
        <div>
          <label>WC / BC *</label>
          <select name="wc_bc" required>
            {% for w in wc_bc %}
            <option value="{{ w }}" {{ 'selected' if record and record.wc_bc==w else '' }}>{{ w }}</option>
            {% endfor %}
          </select>
        </div>
        <div>
          <label>Horas *</label>
          <input type="number" step="0.1" min="0" name="horas" value="{{ record.horas if record else '' }}" required>
        </div>
        <div>
          <label>Tarea *</label>
          <select name="tarea" required>
            {% for t in tareas %}
            <option value="{{ t }}" {{ 'selected' if record and record.tarea==t else '' }}>{{ t }}</option>
            {% endfor %}
          </select>
        </div>
      </div>
      <label>Observaciones</label>
      <textarea name="observaciones" rows="3">{{ record.observaciones if record else '' }}</textarea>
      <div class="form-actions">
        <button class="btn-primary" type="submit">Guardar</button>
        <a class="btn-secondary" href="{{ url_for('vas_list') }}">Cancelar</a>
      </div>
    </form>
  </div>
""" + FOOTER

INV_DASHBOARD_TEMPLATE = BASE_HEAD + NAV + """
  <h1 class="page-title">Inventario de Materiales</h1>
  {% with messages = get_flashed_messages(with_categories=true) %}
    {% for cat, msg in messages %}
      <div class="flash flash-{{ 'error' if cat=='error' else 'success' }}">{{ msg }}</div>
    {% endfor %}
  {% endwith %}

  <div class="kpi-row">
    <div class="kpi-card"><div class="kpi-value">{{ kpi.total_materiales }}</div><div class="kpi-label">Materiales activos</div></div>
    <div class="kpi-card"><div class="kpi-value alert">{{ kpi.bajo_minimo }}</div><div class="kpi-label">Bajo stock mínimo</div></div>
    <div class="kpi-card"><div class="kpi-value">{{ kpi.movs_mes }}</div><div class="kpi-label">Movimientos este mes</div></div>
  </div>

  <div class="card">
    <div class="filters-bar" style="align-items:center;">
      <a class="btn-primary" href="{{ url_for('inv_material_new') }}">Nuevo material</a>
      <a class="btn-primary" href="{{ url_for('inv_mov_new') }}">Registrar movimiento</a>
      <a class="btn-secondary" href="{{ url_for('inv_kardex') }}">Trazabilidad por lote/OT</a>
      <a class="btn-secondary" href="{{ url_for('inv_export') }}">Exportar stock a Excel</a>
    </div>
  </div>

  <div class="table-toolbar" style="margin-bottom:6px;"><h2 style="font-size:15px;color:var(--blue);">Stock actual</h2></div>
  <div class="table-wrap">
    <table class="data-table">
      <thead>
        <tr>
          <th>Código</th><th>Nombre</th><th>Categoría</th><th>Unidad</th>
          <th>Stock actual</th><th>Stock mínimo</th><th>Estado</th><th></th>
        </tr>
      </thead>
      <tbody>
        {% for m in materiales %}
        <tr class="{{ 'row-alert' if m.stock_actual < m.stock_minimo else '' }}">
          <td>{{ m.codigo }}</td>
          <td>{{ m.nombre }}</td>
          <td>{{ m.categoria }}</td>
          <td>{{ m.unidad }}</td>
          <td>{{ '%.2f'|format(m.stock_actual) }}</td>
          <td>{{ '%.2f'|format(m.stock_minimo) }}</td>
          <td>
            {% if m.stock_actual < m.stock_minimo %}
              <span class="pill pill-red">Bajo mínimo</span>
            {% else %}
              <span class="pill pill-green">OK</span>
            {% endif %}
          </td>
          <td class="actions-cell">
            <a class="btn-mini" href="{{ url_for('inv_kardex', material_id=m.id) }}">Ver kardex</a>
            <a class="btn-mini" href="{{ url_for('inv_material_edit', material_id=m.id) }}">Editar</a>
          </td>
        </tr>
        {% else %}
        <tr><td colspan="8" style="text-align:center;color:var(--muted);">Aún no hay materiales cargados.</td></tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
""" + FOOTER

INV_MATERIAL_FORM_TEMPLATE = BASE_HEAD + NAV + """
  <h1 class="page-title">{{ 'Editar material' if material else 'Nuevo material' }}</h1>
  {% with messages = get_flashed_messages(with_categories=true) %}
    {% for cat, msg in messages %}
      <div class="flash flash-{{ 'error' if cat=='error' else 'success' }}">{{ msg }}</div>
    {% endfor %}
  {% endwith %}
  <div class="form-card">
    <form method="post">
      <div class="form-grid">
        <div>
          <label>Código *</label>
          <input type="text" name="codigo" value="{{ material.codigo if material else '' }}" required>
        </div>
        <div>
          <label>Nombre *</label>
          <input type="text" name="nombre" value="{{ material.nombre if material else '' }}" required>
        </div>
        <div>
          <label>Categoría *</label>
          <select name="categoria" required>
            {% for c in categorias %}
            <option value="{{ c }}" {{ 'selected' if material and material.categoria==c else '' }}>{{ c }}</option>
            {% endfor %}
          </select>
        </div>
        <div>
          <label>Unidad de medida *</label>
          <select name="unidad" required>
            {% for u in unidades %}
            <option value="{{ u }}" {{ 'selected' if material and material.unidad==u else '' }}>{{ u }}</option>
            {% endfor %}
          </select>
        </div>
        <div>
          <label>Stock mínimo *</label>
          <input type="number" step="0.01" min="0" name="stock_minimo"
                 value="{{ material.stock_minimo if material else '0' }}" required>
        </div>
      </div>
      <div class="form-actions">
        <button class="btn-primary" type="submit">Guardar</button>
        <a class="btn-secondary" href="{{ url_for('inv_dashboard') }}">Cancelar</a>
      </div>
    </form>
  </div>
""" + FOOTER

INV_MOV_FORM_TEMPLATE = BASE_HEAD + NAV + """
  <h1 class="page-title">Registrar movimiento de inventario</h1>
  {% with messages = get_flashed_messages(with_categories=true) %}
    {% for cat, msg in messages %}
      <div class="flash flash-{{ 'error' if cat=='error' else 'success' }}">{{ msg }}</div>
    {% endfor %}
  {% endwith %}
  <div class="form-card">
    <form method="post">
      <div class="form-grid">
        <div>
          <label>Material *</label>
          <select name="material_id" required>
            <option value="">Selecciona...</option>
            {% for m in materiales %}
            <option value="{{ m.id }}">{{ m.codigo }} - {{ m.nombre }} ({{ m.unidad }})</option>
            {% endfor %}
          </select>
        </div>
        <div>
          <label>Tipo *</label>
          <select name="tipo" required>
            {% for t in tipos %}
            <option value="{{ t }}">{{ t }}</option>
            {% endfor %}
          </select>
        </div>
        <div>
          <label>Cantidad *</label>
          <input type="number" step="0.01" min="0.01" name="cantidad" required>
        </div>
        <div>
          <label>Fecha *</label>
          <input type="date" name="fecha" value="{{ today }}" required>
        </div>
        <div>
          <label>Delivery</label>
          <div style="display:flex;gap:6px;">
            <input type="text" id="delivery-input" name="delivery" placeholder="Ej: 802522080">
            <button type="button" class="btn-secondary" style="white-space:nowrap;" onclick="abrirEscaner('delivery-input')">""" + _icon("scan") + """ Escanear</button>
          </div>
        </div>
        <div>
          <label>Lote / OT</label>
          <div style="display:flex;gap:6px;">
            <input type="text" id="lote-input" name="lote_ot" placeholder="Ej: OT-2026-0451">
            <button type="button" class="btn-secondary" style="white-space:nowrap;" onclick="abrirEscaner('lote-input')">""" + _icon("scan") + """ Escanear</button>
          </div>
        </div>
        <div>
          <label>Responsable</label>
          <input type="text" name="responsable">
        </div>
      </div>
      <label>Observaciones</label>
      <textarea name="observaciones" rows="3"></textarea>
      <div class="form-actions">
        <button class="btn-primary" type="submit">Guardar movimiento</button>
        <a class="btn-secondary" href="{{ url_for('inv_dashboard') }}">Cancelar</a>
      </div>
    </form>
  </div>

  <div id="scan-modal" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.6);z-index:100;align-items:center;justify-content:center;">
    <div style="background:#fff;border-radius:12px;padding:18px;width:min(420px,92vw);">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;">
        <strong style="color:var(--blue);font-size:14px;">Escanear código de barras / QR</strong>
        <button type="button" class="btn-secondary" style="padding:4px 10px;" onclick="cerrarEscaner()">Cerrar</button>
      </div>
      <div id="reader" style="width:100%;"></div>
      <p class="muted" style="margin-top:8px;">Apunta la cámara al código del Delivery o del bulto.</p>
    </div>
  </div>
  <script src="https://cdn.jsdelivr.net/npm/html5-qrcode@2.3.8/html5-qrcode.min.js"></script>
  <script>
    let html5QrCode = null;
    let currentTargetInput = null;

    function abrirEscaner(inputId) {
      currentTargetInput = document.getElementById(inputId);
      document.getElementById('scan-modal').style.display = 'flex';
      html5QrCode = new Html5Qrcode("reader");
      html5QrCode.start(
        { facingMode: "environment" },
        { fps: 10, qrbox: { width: 250, height: 120 } },
        (decodedText) => {
          if (currentTargetInput) { currentTargetInput.value = decodedText; }
          cerrarEscaner();
        },
        () => {}
      ).catch(() => {
        alert('No se pudo acceder a la cámara. Verifica los permisos del navegador.');
        cerrarEscaner();
      });
    }

    function cerrarEscaner() {
      document.getElementById('scan-modal').style.display = 'none';
      if (html5QrCode) {
        html5QrCode.stop().then(() => html5QrCode.clear()).catch(() => {});
      }
    }
  </script>
""" + FOOTER

INV_KARDEX_TEMPLATE = BASE_HEAD + NAV + """
  <h1 class="page-title">Trazabilidad / Kardex</h1>
  <div class="card">
    <form method="get" class="filters-bar">
      <div>
        <label>Material</label>
        <select name="material_id">
          <option value="">Todos</option>
          {% for m in materiales %}
          <option value="{{ m.id }}" {{ 'selected' if filtros.material_id==m.id|string else '' }}>{{ m.codigo }} - {{ m.nombre }}</option>
          {% endfor %}
        </select>
      </div>
      <div>
        <label>Delivery</label>
        <input type="text" name="delivery" value="{{ filtros.delivery or '' }}" placeholder="Ej: 802522080">
      </div>
      <div>
        <label>Lote / OT</label>
        <input type="text" name="lote_ot" value="{{ filtros.lote_ot or '' }}" placeholder="Ej: OT-2026-0451">
      </div>
      <div><button class="btn-secondary" type="submit">Buscar</button></div>
    </form>
  </div>

  <div class="table-wrap">
    <table class="data-table">
      <thead>
        <tr>
          <th>Fecha</th><th>Material</th><th>Tipo</th><th>Cantidad</th>
          <th>Delivery</th><th>Lote/OT</th><th>Responsable</th><th>Observaciones</th>
        </tr>
      </thead>
      <tbody>
        {% for r in movimientos %}
        <tr>
          <td>{{ r.fecha }}</td>
          <td>{{ r.codigo }} - {{ r.nombre }}</td>
          <td>
            {% if r.tipo == 'Entrada' %}
              <span class="pill pill-green">Entrada</span>
            {% else %}
              <span class="pill pill-orange">Salida</span>
            {% endif %}
          </td>
          <td>{{ '%.2f'|format(r.cantidad) }}</td>
          <td>{{ r.delivery or '-' }}</td>
          <td>{{ r.lote_ot or '-' }}</td>
          <td>{{ r.responsable or '-' }}</td>
          <td>{{ r.observaciones or '' }}</td>
        </tr>
        {% else %}
        <tr><td colspan="8" style="text-align:center;color:var(--muted);">Sin movimientos para este filtro.</td></tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
""" + FOOTER

CALIDAD_LIST_TEMPLATE = BASE_HEAD + NAV + """
  <h1 class="page-title">Registro de Calidad</h1>
  {% with messages = get_flashed_messages(with_categories=true) %}
    {% for cat, msg in messages %}
      <div class="flash flash-{{ 'error' if cat=='error' else 'success' }}">{{ msg }}</div>
    {% endfor %}
  {% endwith %}

  <div class="kpi-row">
    <div class="kpi-card"><div class="kpi-value">{{ kpi.total }}</div><div class="kpi-label">Registros (filtro)</div></div>
    <div class="kpi-card"><div class="kpi-value alert">{{ kpi.pendientes }}</div><div class="kpi-label">Pendientes / En proceso</div></div>
    <div class="kpi-card"><div class="kpi-value">{{ kpi.cerrados }}</div><div class="kpi-label">Cerrados / Aprobados</div></div>
    <div class="kpi-card"><div class="kpi-value">{{ kpi.top_tipo }}</div><div class="kpi-label">Tipo más frecuente</div></div>
  </div>

  <div class="card">
    <canvas id="chartTipo" height="90"></canvas>
  </div>

  <div class="card">
    <form method="get" class="filters-bar">
      <div>
        <label>Tipo</label>
        <select name="tipo">
          <option value="">Todos</option>
          {% for t in tipos %}
          <option value="{{ t }}" {{ 'selected' if filtros.tipo==t else '' }}>{{ t }}</option>
          {% endfor %}
        </select>
      </div>
      <div>
        <label>Estado</label>
        <select name="estado">
          <option value="">Todos</option>
          {% for e in estados %}
          <option value="{{ e }}" {{ 'selected' if filtros.estado==e else '' }}>{{ e }}</option>
          {% endfor %}
        </select>
      </div>
      <div>
        <label>Cliente</label>
        <input type="text" name="cliente" value="{{ filtros.cliente or '' }}" placeholder="Ej: Collahuasi">
      </div>
      <div>
        <label>Delivery / Guía / RDEL</label>
        <input type="text" name="buscar" value="{{ filtros.buscar or '' }}" placeholder="Delivery, guía o RDEL">
      </div>
      <div><button class="btn-secondary" type="submit">Filtrar</button></div>
      <div><a class="btn-secondary" href="{{ url_for('calidad_export', **filtros) }}">Exportar Excel</a></div>
      <div><a class="btn-primary" href="{{ url_for('calidad_new') }}">Nuevo registro</a></div>
    </form>
  </div>

  <div class="table-wrap">
    <table class="data-table">
      <thead>
        <tr>
          <th>Fecha</th><th>Tipo</th><th>Delivery</th><th>RDEL</th><th>Cliente</th><th>Ítem</th>
          <th>Cantidad</th><th>Motivo</th><th>Responsable</th><th>Estado</th><th></th>
        </tr>
      </thead>
      <tbody>
        {% for r in registros %}
        <tr>
          <td>{{ r.fecha }}</td>
          <td><span class="pill pill-blue">{{ r.tipo }}</span></td>
          <td>{{ r.delivery or '-' }}</td>
          <td>{{ r.rdel or '-' }}</td>
          <td>{{ r.cliente or '-' }}</td>
          <td>{{ r.item or '-' }}</td>
          <td>{{ r.cantidad if r.cantidad is not none else '-' }}</td>
          <td>{{ (r.motivo or '')[:60] }}{{ '…' if r.motivo and r.motivo|length > 60 else '' }}</td>
          <td>{{ r.responsable or '-' }}</td>
          <td>
            {% if r.estado in ['Cerrado', 'Aprobado'] %}
              <span class="pill pill-green">{{ r.estado }}</span>
            {% elif r.estado == 'Rechazado' %}
              <span class="pill pill-red">{{ r.estado }}</span>
            {% elif r.estado == 'Esperando RDEL' %}
              <span class="pill pill-orange">{{ r.estado }}</span>
            {% else %}
              <span class="pill pill-orange">{{ r.estado }}</span>
            {% endif %}
          </td>
          <td class="actions-cell">
            <a class="btn-mini" href="{{ url_for('calidad_edit', record_id=r.id) }}">Editar</a>
            <form method="post" action="{{ url_for('calidad_delete', record_id=r.id) }}" style="display:inline;"
                  onsubmit="return confirm('¿Eliminar este registro?');">
              <button class="btn-mini btn-mini-danger" type="submit">Eliminar</button>
            </form>
          </td>
        </tr>
        {% else %}
        <tr><td colspan="11" style="text-align:center;color:var(--muted);">Sin registros para este filtro.</td></tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
<script>
  const ctxTipo = document.getElementById('chartTipo');
  new Chart(ctxTipo, {
    type: 'bar',
    data: {
      labels: {{ chart_labels|tojson }},
      datasets: [{
        label: 'Registros por Tipo',
        data: {{ chart_data|tojson }},
        backgroundColor: '#144E8C'
      }]
    },
    options: { responsive:true, plugins:{legend:{display:false}} }
  });
</script>
""" + FOOTER

CALIDAD_FORM_TEMPLATE = BASE_HEAD + NAV + """
  <h1 class="page-title">{{ 'Editar registro de Calidad' if record else 'Nuevo registro de Calidad' }}</h1>
  {% with messages = get_flashed_messages(with_categories=true) %}
    {% for cat, msg in messages %}
      <div class="flash flash-{{ 'error' if cat=='error' else 'success' }}">{{ msg }}</div>
    {% endfor %}
  {% endwith %}
  <div class="form-card form-card-wide">
    <form method="post">
      <div class="form-grid">
        <div>
          <label>Tipo *</label>
          <select name="tipo" id="tipo-select" required onchange="actualizarCamposTipo()">
            {% for t in tipos %}
            <option value="{{ t }}" {{ 'selected' if record and record.tipo==t else '' }}>{{ t }}</option>
            {% endfor %}
          </select>
        </div>
        <div>
          <label>Fecha *</label>
          <input type="date" name="fecha" value="{{ record.fecha if record else today }}" required>
        </div>
        <div>
          <label>Estado *</label>
          <select name="estado" required>
            {% for e in estados %}
            <option value="{{ e }}" {{ 'selected' if record and record.estado==e else '' }}>{{ e }}</option>
            {% endfor %}
          </select>
        </div>
        <div>
          <label>Delivery</label>
          <div style="display:flex;gap:6px;">
            <input type="text" id="delivery-input-cal" name="delivery" value="{{ record.delivery if record else '' }}">
            <button type="button" class="btn-secondary" style="white-space:nowrap;" onclick="abrirEscanerCal('delivery-input-cal')">""" + _icon("scan") + """ Escanear</button>
          </div>
        </div>
        <div>
          <label>Guía</label>
          <input type="text" name="guia" value="{{ record.guia if record else '' }}">
        </div>
        <div>
          <label>Cliente</label>
          <input type="text" name="cliente" value="{{ record.cliente if record else '' }}">
        </div>
      </div>

      <div class="form-section-label">Material</div>
      {% if record %}
      <div class="form-grid">
        <div>
          <label>Ítem / PN</label>
          <input type="text" name="item" value="{{ record.item if record else '' }}">
        </div>
        <div>
          <label>Descripción material</label>
          <input type="text" name="descripcion_material" value="{{ record.descripcion_material if record else '' }}">
        </div>
        <div>
          <label>Cantidad</label>
          <input type="number" step="0.01" name="cantidad" value="{{ record.cantidad if record else '' }}">
        </div>
      </div>
      <p class="muted">Estás editando un registro existente, que corresponde a un solo ítem. Si este caso tiene más ítems, agrégalos como registros nuevos (usando el mismo Delivery/Guía) desde "Nuevo registro".</p>
      {% else %}
      <p class="muted" style="margin-top:-4px;">Si el Delivery/Guía trae más de un ítem, agrega una fila por cada uno — se creará un registro por ítem, todos con los mismos datos generales de arriba.</p>
      <div class="table-wrap" style="margin-bottom:10px;">
        <table class="data-table" id="items-table">
          <thead>
            <tr><th style="width:30%;">Ítem / PN</th><th>Descripción material</th><th style="width:15%;">Cantidad</th><th></th></tr>
          </thead>
          <tbody id="items-tbody">
            <tr>
              <td><input type="text" name="item[]"></td>
              <td><input type="text" name="descripcion_material[]"></td>
              <td><input type="number" step="0.01" name="cantidad[]"></td>
              <td><button type="button" class="btn-mini btn-mini-danger" onclick="quitarFilaItem(this)">Quitar</button></td>
            </tr>
          </tbody>
        </table>
      </div>
      <button type="button" class="btn-secondary" onclick="agregarFilaItem()">+ Agregar ítem</button>
      {% endif %}

      <div class="form-section-label">Datos adicionales</div>
      <p class="muted" style="margin-top:-4px;">Ubicación y Peso aplican principalmente a Scrap y Reversa; Proceso es el estado físico del material (ej: "Esperando retiro de scrap"). RDEL es el N° de retorno que asigna el mandante (Metso) para poder ingresar el Delivery retornado a inventario.</p>
      <div class="form-grid">
        <div>
          <label>Ubicación</label>
          <input type="text" name="ubicacion" value="{{ record.ubicacion if record else '' }}" placeholder="Ej: AO-B-91-005-000">
        </div>
        <div>
          <label>Peso (kg)</label>
          <input type="number" step="0.01" name="peso_kg" value="{{ record.peso_kg if record else '' }}">
        </div>
        <div>
          <label>Proceso</label>
          <input type="text" name="proceso" value="{{ record.proceso if record else '' }}" placeholder="Ej: Esperando retiro de scrap">
        </div>
        <div>
          <label>RDEL</label>
          <input type="text" name="rdel" value="{{ record.rdel if record else '' }}" placeholder="N° asignado por Metso">
        </div>
      </div>

      <div class="form-section-label" id="seccion-diferencia-stock" style="display:none;">
        Diferencia de Stock (solo aplica a este tipo)
      </div>
      <div class="form-grid" id="campos-diferencia-stock" style="display:none;">
        <div>
          <label>Cantidad sistema</label>
          <input type="number" step="0.01" name="cantidad_sistema" value="{{ record.cantidad_sistema if record else '' }}">
        </div>
        <div>
          <label>Cantidad física</label>
          <input type="number" step="0.01" name="cantidad_fisica" value="{{ record.cantidad_fisica if record else '' }}">
        </div>
      </div>

      <div class="form-section-label">Análisis</div>
      <label>Motivo / Descripción del caso</label>
      <textarea name="motivo" rows="2">{{ record.motivo if record else '' }}</textarea>
      <label>Causa raíz</label>
      <textarea name="causa_raiz" rows="2">{{ record.causa_raiz if record else '' }}</textarea>
      <label>Acciones correctivas</label>
      <textarea name="acciones_correctivas" rows="2">{{ record.acciones_correctivas if record else '' }}</textarea>

      <div class="form-grid" style="margin-top:10px;">
        <div>
          <label>Responsable</label>
          <input type="text" name="responsable" value="{{ record.responsable if record else '' }}">
        </div>
      </div>
      <label>Observaciones</label>
      <textarea name="observaciones" rows="2">{{ record.observaciones if record else '' }}</textarea>

      <div class="form-actions">
        <button class="btn-primary" type="submit">Guardar</button>
        <a class="btn-secondary" href="{{ url_for('calidad_list') }}">Cancelar</a>
      </div>
    </form>
  </div>

  <div id="scan-modal-cal" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.6);z-index:100;align-items:center;justify-content:center;">
    <div style="background:#fff;border-radius:12px;padding:18px;width:min(420px,92vw);">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;">
        <strong style="color:var(--blue);font-size:14px;">Escanear código de barras / QR</strong>
        <button type="button" class="btn-secondary" style="padding:4px 10px;" onclick="cerrarEscanerCal()">Cerrar</button>
      </div>
      <div id="reader-cal" style="width:100%;"></div>
      <p class="muted" style="margin-top:8px;">Apunta la cámara al código del Delivery o del bulto.</p>
    </div>
  </div>
  <script src="https://cdn.jsdelivr.net/npm/html5-qrcode@2.3.8/html5-qrcode.min.js"></script>
  <script>
    let html5QrCodeCal = null;
    let currentTargetInputCal = null;

    function abrirEscanerCal(inputId) {
      currentTargetInputCal = document.getElementById(inputId);
      document.getElementById('scan-modal-cal').style.display = 'flex';
      html5QrCodeCal = new Html5Qrcode("reader-cal");
      html5QrCodeCal.start(
        { facingMode: "environment" },
        { fps: 10, qrbox: { width: 250, height: 120 } },
        (decodedText) => {
          if (currentTargetInputCal) { currentTargetInputCal.value = decodedText; }
          cerrarEscanerCal();
        },
        () => {}
      ).catch(() => {
        alert('No se pudo acceder a la cámara. Verifica los permisos del navegador.');
        cerrarEscanerCal();
      });
    }

    function cerrarEscanerCal() {
      document.getElementById('scan-modal-cal').style.display = 'none';
      if (html5QrCodeCal) {
        html5QrCodeCal.stop().then(() => html5QrCodeCal.clear()).catch(() => {});
      }
    }

    function actualizarCamposTipo() {
      const tipo = document.getElementById('tipo-select').value;
      const esDiferencia = tipo === 'Diferencia de Stock';
      document.getElementById('seccion-diferencia-stock').style.display = esDiferencia ? 'block' : 'none';
      document.getElementById('campos-diferencia-stock').style.display = esDiferencia ? 'grid' : 'none';
    }
    actualizarCamposTipo();

    function agregarFilaItem() {
      const tbody = document.getElementById('items-tbody');
      if (!tbody) return;
      const fila = document.createElement('tr');
      fila.innerHTML = `
        <td><input type="text" name="item[]"></td>
        <td><input type="text" name="descripcion_material[]"></td>
        <td><input type="number" step="0.01" name="cantidad[]"></td>
        <td><button type="button" class="btn-mini btn-mini-danger" onclick="quitarFilaItem(this)">Quitar</button></td>
      `;
      tbody.appendChild(fila);
    }

    function quitarFilaItem(btn) {
      const tbody = document.getElementById('items-tbody');
      if (tbody && tbody.rows.length > 1) {
        btn.closest('tr').remove();
      }
    }
  </script>
""" + FOOTER

ADMIN_USERS_TEMPLATE = BASE_HEAD + NAV + """
  <h1 class="page-title">Usuarios</h1>
  {% with messages = get_flashed_messages(with_categories=true) %}
    {% for cat, msg in messages %}
      <div class="flash flash-{{ 'error' if cat=='error' else 'success' }}">{{ msg }}</div>
    {% endfor %}
  {% endwith %}

  <div class="form-card" style="margin-bottom:20px;">
    <h2 style="font-size:15px;color:var(--blue);margin:0 0 6px;">Crear usuario</h2>
    <form method="post" action="{{ url_for('admin_users_new') }}">
      <div class="form-grid">
        <div>
          <label>Usuario *</label>
          <input type="text" name="username" required>
        </div>
        <div>
          <label>Contraseña *</label>
          <input type="password" name="password" required minlength="4">
        </div>
        <div>
          <label>Rol *</label>
          <select name="role" required>
            <option value="user">Usuario</option>
            <option value="admin">Administrador</option>
          </select>
        </div>
      </div>
      <div class="form-actions">
        <button class="btn-primary" type="submit">Crear usuario</button>
      </div>
    </form>
  </div>

  <div class="table-wrap">
    <table class="data-table">
      <thead>
        <tr><th>Usuario</th><th>Rol</th><th>Estado</th><th>Creado</th><th></th></tr>
      </thead>
      <tbody>
        {% for u in usuarios %}
        <tr>
          <td>{{ u.username }}</td>
          <td>
            {% if u.role == 'admin' %}
              <span class="pill pill-blue">Administrador</span>
            {% else %}
              <span class="pill pill-orange">Usuario</span>
            {% endif %}
          </td>
          <td>
            {% if u.activo %}
              <span class="pill pill-green">Activo</span>
            {% else %}
              <span class="pill pill-red">Desactivado</span>
            {% endif %}
          </td>
          <td>{{ u.created_at }}</td>
          <td class="actions-cell">
            {% if u.id != session.get('user_id') %}
            <form method="post" action="{{ url_for('admin_users_toggle', user_id=u.id) }}" style="display:inline;">
              <button class="btn-mini {{ 'btn-mini-danger' if u.activo else '' }}" type="submit">
                {{ 'Desactivar' if u.activo else 'Activar' }}
              </button>
            </form>
            {% else %}
            <span class="muted">(tu cuenta)</span>
            {% endif %}
          </td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
""" + FOOTER


# --------------------------------------------------------------------------
# Routes: Auth
# --------------------------------------------------------------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        db = get_db()
        user = db.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
        if user and not user["activo"]:
            flash("Este usuario está desactivado. Contacta a un administrador.", "error")
        elif user and check_password_hash(user["password_hash"], password):
            session["user_id"] = user["id"]
            session["username"] = user["username"]
            session["role"] = user["role"]
            next_url = request.args.get("next") or url_for("hub")
            return redirect(next_url)
        else:
            flash("Usuario o contraseña incorrectos.", "error")
    return render_template_string(LOGIN_TEMPLATE, title="Ingresar")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/logo.png")
def logo_png():
    return Response(base64.b64decode(NEFAB_LOGO_B64), mimetype="image/png")


# --------------------------------------------------------------------------
# Routes: Hub
# --------------------------------------------------------------------------
@app.route("/")
@login_required
def hub():
    return render_template_string(HUB_TEMPLATE, title="Inicio", active="hub")


# --------------------------------------------------------------------------
# Routes: VAS
# --------------------------------------------------------------------------
def _vas_filtered_query():
    mes = request.args.get("mes", "")
    area = request.args.get("area", "")
    tarea = request.args.get("tarea", "")

    query = "SELECT * FROM vas_records WHERE 1=1"
    params = []
    if mes:
        query += " AND strftime('%Y-%m', fecha) = ?"
        params.append(mes)
    if area:
        query += " AND area = ?"
        params.append(area)
    if tarea:
        query += " AND tarea = ?"
        params.append(tarea)
    query += " ORDER BY fecha DESC, id DESC"
    return query, params, {"mes": mes, "area": area, "tarea": tarea}


@app.route("/vas")
@login_required
def vas_list():
    db = get_db()
    query, params, filtros = _vas_filtered_query()
    registros = db.execute(query, params).fetchall()

    total_horas = sum(r["horas"] for r in registros)
    total_registros = len(registros)

    horas_por_area = {}
    horas_por_tarea = {}
    for r in registros:
        horas_por_area[r["area"]] = horas_por_area.get(r["area"], 0) + r["horas"]
        horas_por_tarea[r["tarea"]] = horas_por_tarea.get(r["tarea"], 0) + r["horas"]

    top_area = max(horas_por_area, key=horas_por_area.get) if horas_por_area else "-"
    top_tarea = max(horas_por_tarea, key=horas_por_tarea.get) if horas_por_tarea else "-"

    kpi = {
        "total_horas": f"{total_horas:.1f}",
        "total_registros": total_registros,
        "top_area": top_area,
        "top_tarea": top_tarea,
    }

    return render_template_string(
        VAS_LIST_TEMPLATE, title="VAS",
        registros=registros, kpi=kpi, filtros=filtros,
        areas=AREAS_VAS, tareas=TAREAS_VAS,
        chart_labels=list(horas_por_area.keys()),
        chart_data=list(horas_por_area.values()),
        active="vas",
    )


@app.route("/vas/nuevo", methods=["GET", "POST"])
@login_required
def vas_new():
    if request.method == "POST":
        db = get_db()
        db.execute("""
            INSERT INTO vas_records (delivery, area, fecha, nombre, wc_bc, horas, tarea, observaciones, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            request.form.get("delivery", "").strip(),
            request.form["area"],
            request.form["fecha"],
            request.form["nombre"].strip(),
            request.form["wc_bc"],
            float(request.form["horas"] or 0),
            request.form["tarea"],
            request.form.get("observaciones", "").strip(),
            session.get("username"),
        ))
        db.commit()
        flash("Registro creado correctamente.", "success")
        return redirect(url_for("vas_list"))

    return render_template_string(
        VAS_FORM_TEMPLATE, title="Nuevo registro VAS",
        record=None, areas=AREAS_VAS, wc_bc=WC_BC_OPTIONS, tareas=TAREAS_VAS,
        today=date.today().isoformat(), active="vas"
    )


@app.route("/vas/editar/<int:record_id>", methods=["GET", "POST"])
@login_required
def vas_edit(record_id):
    db = get_db()
    record = db.execute("SELECT * FROM vas_records WHERE id = ?", (record_id,)).fetchone()
    if record is None:
        flash("Registro no encontrado.", "error")
        return redirect(url_for("vas_list"))

    if request.method == "POST":
        db.execute("""
            UPDATE vas_records
            SET delivery=?, area=?, fecha=?, nombre=?, wc_bc=?, horas=?, tarea=?, observaciones=?,
                updated_at=datetime('now')
            WHERE id=?
        """, (
            request.form.get("delivery", "").strip(),
            request.form["area"],
            request.form["fecha"],
            request.form["nombre"].strip(),
            request.form["wc_bc"],
            float(request.form["horas"] or 0),
            request.form["tarea"],
            request.form.get("observaciones", "").strip(),
            record_id,
        ))
        db.commit()
        flash("Registro actualizado.", "success")
        return redirect(url_for("vas_list"))

    return render_template_string(
        VAS_FORM_TEMPLATE, title="Editar registro VAS",
        record=record, areas=AREAS_VAS, wc_bc=WC_BC_OPTIONS, tareas=TAREAS_VAS,
        today=date.today().isoformat(), active="vas"
    )


@app.route("/vas/eliminar/<int:record_id>", methods=["POST"])
@login_required
def vas_delete(record_id):
    db = get_db()
    db.execute("DELETE FROM vas_records WHERE id = ?", (record_id,))
    db.commit()
    flash("Registro eliminado.", "success")
    return redirect(url_for("vas_list"))


@app.route("/vas/exportar")
@login_required
def vas_export():
    db = get_db()
    query, params, _ = _vas_filtered_query()
    registros = db.execute(query, params).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VAS"

    headers = ["Fecha", "Delivery", "Área", "Nombre", "WC/BC", "Horas", "Tarea", "Observaciones"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = PatternFill("solid", fgColor="144E8C")
        cell.alignment = Alignment(horizontal="center")

    for r in registros:
        ws.append([r["fecha"], r["delivery"], r["area"], r["nombre"], r["wc_bc"],
                   r["horas"], r["tarea"], r["observaciones"]])

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"VAS_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# --------------------------------------------------------------------------
# Routes: Inventario
# --------------------------------------------------------------------------
def _stock_actual(db, material_id):
    row = db.execute("""
        SELECT
          COALESCE(SUM(CASE WHEN tipo='Entrada' THEN cantidad ELSE 0 END), 0) -
          COALESCE(SUM(CASE WHEN tipo='Salida' THEN cantidad ELSE 0 END), 0) AS stock
        FROM movimientos_inventario WHERE material_id = ?
    """, (material_id,)).fetchone()
    return row["stock"] or 0


@app.route("/inventario")
@login_required
def inv_dashboard():
    db = get_db()
    materiales_raw = db.execute(
        "SELECT * FROM materiales WHERE activo = 1 ORDER BY categoria, nombre"
    ).fetchall()

    materiales = []
    bajo_minimo = 0
    for m in materiales_raw:
        stock = _stock_actual(db, m["id"])
        d = dict(m)
        d["stock_actual"] = stock
        if stock < m["stock_minimo"]:
            bajo_minimo += 1
        materiales.append(d)

    movs_mes = db.execute(
        "SELECT COUNT(*) AS c FROM movimientos_inventario WHERE strftime('%Y-%m', fecha) = strftime('%Y-%m', 'now')"
    ).fetchone()["c"]

    kpi = {
        "total_materiales": len(materiales),
        "bajo_minimo": bajo_minimo,
        "movs_mes": movs_mes,
    }

    return render_template_string(INV_DASHBOARD_TEMPLATE, title="Inventario",
                                   materiales=materiales, kpi=kpi, active="inv")


@app.route("/inventario/material/nuevo", methods=["GET", "POST"])
@login_required
def inv_material_new():
    if request.method == "POST":
        db = get_db()
        try:
            db.execute("""
                INSERT INTO materiales (codigo, nombre, categoria, unidad, stock_minimo)
                VALUES (?, ?, ?, ?, ?)
            """, (
                request.form["codigo"].strip(),
                request.form["nombre"].strip(),
                request.form["categoria"],
                request.form["unidad"],
                float(request.form["stock_minimo"] or 0),
            ))
            db.commit()
            flash("Material creado correctamente.", "success")
            return redirect(url_for("inv_dashboard"))
        except sqlite3.IntegrityError:
            flash("Ya existe un material con ese código.", "error")

    return render_template_string(INV_MATERIAL_FORM_TEMPLATE, title="Nuevo material",
                                   material=None, categorias=CATEGORIAS_MATERIAL,
                                   unidades=UNIDADES_MEDIDA, active="inv")


@app.route("/inventario/material/editar/<int:material_id>", methods=["GET", "POST"])
@login_required
def inv_material_edit(material_id):
    db = get_db()
    material = db.execute("SELECT * FROM materiales WHERE id = ?", (material_id,)).fetchone()
    if material is None:
        flash("Material no encontrado.", "error")
        return redirect(url_for("inv_dashboard"))

    if request.method == "POST":
        try:
            db.execute("""
                UPDATE materiales SET codigo=?, nombre=?, categoria=?, unidad=?, stock_minimo=?
                WHERE id=?
            """, (
                request.form["codigo"].strip(),
                request.form["nombre"].strip(),
                request.form["categoria"],
                request.form["unidad"],
                float(request.form["stock_minimo"] or 0),
                material_id,
            ))
            db.commit()
            flash("Material actualizado.", "success")
            return redirect(url_for("inv_dashboard"))
        except sqlite3.IntegrityError:
            flash("Ya existe un material con ese código.", "error")

    return render_template_string(INV_MATERIAL_FORM_TEMPLATE, title="Editar material",
                                   material=material, categorias=CATEGORIAS_MATERIAL,
                                   unidades=UNIDADES_MEDIDA, active="inv")


@app.route("/inventario/movimiento/nuevo", methods=["GET", "POST"])
@login_required
def inv_mov_new():
    db = get_db()
    materiales = db.execute(
        "SELECT * FROM materiales WHERE activo = 1 ORDER BY categoria, nombre"
    ).fetchall()

    if request.method == "POST":
        material_id = request.form.get("material_id")
        if not material_id:
            flash("Debes seleccionar un material.", "error")
        else:
            db.execute("""
                INSERT INTO movimientos_inventario
                    (material_id, tipo, cantidad, lote_ot, delivery, fecha, responsable, observaciones, created_by)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                material_id,
                request.form["tipo"],
                float(request.form["cantidad"]),
                request.form.get("lote_ot", "").strip(),
                request.form.get("delivery", "").strip(),
                request.form["fecha"],
                request.form.get("responsable", "").strip(),
                request.form.get("observaciones", "").strip(),
                session.get("username"),
            ))
            db.commit()
            flash("Movimiento registrado.", "success")
            return redirect(url_for("inv_dashboard"))

    return render_template_string(INV_MOV_FORM_TEMPLATE, title="Nuevo movimiento",
                                   materiales=materiales, tipos=TIPOS_MOVIMIENTO,
                                   today=date.today().isoformat(), active="inv")


@app.route("/inventario/kardex")
@login_required
def inv_kardex():
    db = get_db()
    material_id = request.args.get("material_id", "")
    lote_ot = request.args.get("lote_ot", "")
    delivery = request.args.get("delivery", "")

    query = """
        SELECT mv.*, ma.codigo, ma.nombre
        FROM movimientos_inventario mv
        JOIN materiales ma ON ma.id = mv.material_id
        WHERE 1=1
    """
    params = []
    if material_id:
        query += " AND mv.material_id = ?"
        params.append(material_id)
    if lote_ot:
        query += " AND mv.lote_ot LIKE ?"
        params.append(f"%{lote_ot}%")
    if delivery:
        query += " AND mv.delivery LIKE ?"
        params.append(f"%{delivery}%")
    query += " ORDER BY mv.fecha DESC, mv.id DESC"

    movimientos = db.execute(query, params).fetchall()
    materiales = db.execute("SELECT * FROM materiales ORDER BY categoria, nombre").fetchall()

    return render_template_string(INV_KARDEX_TEMPLATE, title="Kardex",
                                   movimientos=movimientos, materiales=materiales,
                                   filtros={"material_id": material_id, "lote_ot": lote_ot, "delivery": delivery},
                                   active="inv")


@app.route("/inventario/exportar")
@login_required
def inv_export():
    db = get_db()
    materiales_raw = db.execute(
        "SELECT * FROM materiales WHERE activo = 1 ORDER BY categoria, nombre"
    ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock"

    headers = ["Código", "Nombre", "Categoría", "Unidad", "Stock actual", "Stock mínimo", "Estado"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = PatternFill("solid", fgColor="144E8C")
        cell.alignment = Alignment(horizontal="center")

    for m in materiales_raw:
        stock = _stock_actual(db, m["id"])
        estado = "Bajo mínimo" if stock < m["stock_minimo"] else "OK"
        ws.append([m["codigo"], m["nombre"], m["categoria"], m["unidad"],
                   stock, m["stock_minimo"], estado])

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Inventario_stock_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# --------------------------------------------------------------------------
# Routes: Calidad
# --------------------------------------------------------------------------
def _calidad_filtered_query():
    tipo = request.args.get("tipo", "")
    estado = request.args.get("estado", "")
    cliente = request.args.get("cliente", "")
    buscar = request.args.get("buscar", "")

    query = "SELECT * FROM calidad_registros WHERE 1=1"
    params = []
    if tipo:
        query += " AND tipo = ?"
        params.append(tipo)
    if estado:
        query += " AND estado = ?"
        params.append(estado)
    if cliente:
        query += " AND cliente LIKE ?"
        params.append(f"%{cliente}%")
    if buscar:
        query += " AND (delivery LIKE ? OR guia LIKE ? OR rdel LIKE ?)"
        params.append(f"%{buscar}%")
        params.append(f"%{buscar}%")
        params.append(f"%{buscar}%")
    query += " ORDER BY fecha DESC, id DESC"
    return query, params, {"tipo": tipo, "estado": estado, "cliente": cliente, "buscar": buscar}


@app.route("/calidad")
@login_required
def calidad_list():
    db = get_db()
    query, params, filtros = _calidad_filtered_query()
    registros = db.execute(query, params).fetchall()

    total = len(registros)
    pendientes = sum(1 for r in registros if r["estado"] in ("Pendiente", "En proceso"))
    cerrados = sum(1 for r in registros if r["estado"] in ("Cerrado", "Aprobado"))

    por_tipo = {}
    for r in registros:
        por_tipo[r["tipo"]] = por_tipo.get(r["tipo"], 0) + 1
    top_tipo = max(por_tipo, key=por_tipo.get) if por_tipo else "-"

    kpi = {"total": total, "pendientes": pendientes, "cerrados": cerrados, "top_tipo": top_tipo}

    return render_template_string(
        CALIDAD_LIST_TEMPLATE, title="Calidad",
        registros=registros, kpi=kpi, filtros=filtros,
        tipos=TIPOS_CALIDAD, estados=ESTADOS_CALIDAD,
        chart_labels=list(por_tipo.keys()), chart_data=list(por_tipo.values()),
        active="calidad",
    )


def _calidad_form_values():
    def to_float(key):
        val = request.form.get(key, "").strip()
        return float(val) if val else None

    return (
        request.form["tipo"],
        request.form["fecha"],
        request.form.get("guia", "").strip(),
        request.form.get("delivery", "").strip(),
        request.form.get("cliente", "").strip(),
        request.form.get("item", "").strip(),
        request.form.get("descripcion_material", "").strip(),
        to_float("cantidad"),
        request.form.get("ubicacion", "").strip(),
        to_float("cantidad_sistema"),
        to_float("cantidad_fisica"),
        to_float("peso_kg"),
        request.form.get("proceso", "").strip(),
        request.form.get("rdel", "").strip(),
        request.form.get("motivo", "").strip(),
        request.form.get("causa_raiz", "").strip(),
        request.form.get("acciones_correctivas", "").strip(),
        request.form.get("responsable", "").strip(),
        request.form["estado"],
        request.form.get("observaciones", "").strip(),
    )


def _calidad_header_values():
    """Campos comunes a todos los ítems de un mismo caso (todo excepto item/descripcion/cantidad)."""
    def to_float(key):
        val = request.form.get(key, "").strip()
        return float(val) if val else None

    return {
        "tipo": request.form["tipo"],
        "fecha": request.form["fecha"],
        "guia": request.form.get("guia", "").strip(),
        "delivery": request.form.get("delivery", "").strip(),
        "cliente": request.form.get("cliente", "").strip(),
        "ubicacion": request.form.get("ubicacion", "").strip(),
        "cantidad_sistema": to_float("cantidad_sistema"),
        "cantidad_fisica": to_float("cantidad_fisica"),
        "peso_kg": to_float("peso_kg"),
        "proceso": request.form.get("proceso", "").strip(),
        "rdel": request.form.get("rdel", "").strip(),
        "motivo": request.form.get("motivo", "").strip(),
        "causa_raiz": request.form.get("causa_raiz", "").strip(),
        "acciones_correctivas": request.form.get("acciones_correctivas", "").strip(),
        "responsable": request.form.get("responsable", "").strip(),
        "estado": request.form["estado"],
        "observaciones": request.form.get("observaciones", "").strip(),
    }


@app.route("/calidad/nuevo", methods=["GET", "POST"])
@login_required
def calidad_new():
    if request.method == "POST":
        db = get_db()
        header = _calidad_header_values()

        items = request.form.getlist("item[]")
        descripciones = request.form.getlist("descripcion_material[]")
        cantidades = request.form.getlist("cantidad[]")

        filas_creadas = 0
        for i in range(len(items)):
            item = items[i].strip()
            descripcion = descripciones[i].strip() if i < len(descripciones) else ""
            cantidad_raw = cantidades[i].strip() if i < len(cantidades) else ""
            if not item and not descripcion and not cantidad_raw:
                continue  # fila vacia, se omite
            cantidad = float(cantidad_raw) if cantidad_raw else None

            db.execute("""
                INSERT INTO calidad_registros
                    (tipo, fecha, guia, delivery, cliente, item, descripcion_material, cantidad,
                     ubicacion, cantidad_sistema, cantidad_fisica, peso_kg, proceso, rdel, motivo, causa_raiz,
                     acciones_correctivas, responsable, estado, observaciones, created_by)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                header["tipo"], header["fecha"], header["guia"], header["delivery"], header["cliente"],
                item, descripcion, cantidad,
                header["ubicacion"], header["cantidad_sistema"], header["cantidad_fisica"],
                header["peso_kg"], header["proceso"], header["rdel"],
                header["motivo"], header["causa_raiz"], header["acciones_correctivas"],
                header["responsable"], header["estado"], header["observaciones"],
                session.get("username"),
            ))
            filas_creadas += 1

        if filas_creadas == 0:
            # Sin ninguna fila de item con datos: igual guarda un registro con el caso general
            db.execute("""
                INSERT INTO calidad_registros
                    (tipo, fecha, guia, delivery, cliente, item, descripcion_material, cantidad,
                     ubicacion, cantidad_sistema, cantidad_fisica, peso_kg, proceso, rdel, motivo, causa_raiz,
                     acciones_correctivas, responsable, estado, observaciones, created_by)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                header["tipo"], header["fecha"], header["guia"], header["delivery"], header["cliente"],
                "", "", None,
                header["ubicacion"], header["cantidad_sistema"], header["cantidad_fisica"],
                header["peso_kg"], header["proceso"], header["rdel"],
                header["motivo"], header["causa_raiz"], header["acciones_correctivas"],
                header["responsable"], header["estado"], header["observaciones"],
                session.get("username"),
            ))
            filas_creadas = 1

        db.commit()
        if filas_creadas > 1:
            flash(f"{filas_creadas} registros creados correctamente (uno por ítem).", "success")
        else:
            flash("Registro de Calidad creado correctamente.", "success")
        return redirect(url_for("calidad_list"))

    return render_template_string(
        CALIDAD_FORM_TEMPLATE, title="Nuevo registro de Calidad",
        record=None, tipos=TIPOS_CALIDAD, estados=ESTADOS_CALIDAD,
        today=date.today().isoformat(), active="calidad",
    )


@app.route("/calidad/editar/<int:record_id>", methods=["GET", "POST"])
@login_required
def calidad_edit(record_id):
    db = get_db()
    record = db.execute("SELECT * FROM calidad_registros WHERE id = ?", (record_id,)).fetchone()
    if record is None:
        flash("Registro no encontrado.", "error")
        return redirect(url_for("calidad_list"))

    if request.method == "POST":
        values = _calidad_form_values()
        db.execute("""
            UPDATE calidad_registros
            SET tipo=?, fecha=?, guia=?, delivery=?, cliente=?, item=?, descripcion_material=?,
                cantidad=?, ubicacion=?, cantidad_sistema=?, cantidad_fisica=?, peso_kg=?, proceso=?, rdel=?,
                motivo=?, causa_raiz=?, acciones_correctivas=?, responsable=?, estado=?, observaciones=?,
                updated_at=datetime('now')
            WHERE id=?
        """, values + (record_id,))
        db.commit()
        flash("Registro actualizado.", "success")
        return redirect(url_for("calidad_list"))

    return render_template_string(
        CALIDAD_FORM_TEMPLATE, title="Editar registro de Calidad",
        record=record, tipos=TIPOS_CALIDAD, estados=ESTADOS_CALIDAD,
        today=date.today().isoformat(), active="calidad",
    )


@app.route("/calidad/eliminar/<int:record_id>", methods=["POST"])
@login_required
def calidad_delete(record_id):
    db = get_db()
    db.execute("DELETE FROM calidad_registros WHERE id = ?", (record_id,))
    db.commit()
    flash("Registro eliminado.", "success")
    return redirect(url_for("calidad_list"))


@app.route("/calidad/exportar")
@login_required
def calidad_export():
    db = get_db()
    query, params, _ = _calidad_filtered_query()
    registros = db.execute(query, params).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calidad"

    headers = ["Fecha", "Tipo", "Guía", "Delivery", "RDEL", "Cliente", "Ítem", "Descripción material",
               "Cantidad", "Ubicación", "Cant. sistema", "Cant. física", "Peso (kg)", "Proceso",
               "Motivo", "Causa raíz", "Acciones correctivas", "Responsable", "Estado", "Observaciones"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = PatternFill("solid", fgColor="144E8C")
        cell.alignment = Alignment(horizontal="center")

    for r in registros:
        ws.append([r["fecha"], r["tipo"], r["guia"], r["delivery"], r["rdel"], r["cliente"], r["item"],
                   r["descripcion_material"], r["cantidad"], r["ubicacion"], r["cantidad_sistema"],
                   r["cantidad_fisica"], r["peso_kg"], r["proceso"], r["motivo"], r["causa_raiz"],
                   r["acciones_correctivas"], r["responsable"], r["estado"], r["observaciones"]])

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Calidad_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# --------------------------------------------------------------------------
# Routes: Admin - Usuarios
# --------------------------------------------------------------------------
@app.route("/admin/usuarios")
@admin_required
def admin_users():
    db = get_db()
    usuarios = db.execute("SELECT * FROM users ORDER BY created_at").fetchall()
    return render_template_string(ADMIN_USERS_TEMPLATE, title="Usuarios",
                                   usuarios=usuarios, active="users")


@app.route("/admin/usuarios/nuevo", methods=["POST"])
@admin_required
def admin_users_new():
    db = get_db()
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")
    role = request.form.get("role", "user")

    if not username or not password:
        flash("Usuario y contraseña son obligatorios.", "error")
        return redirect(url_for("admin_users"))

    try:
        db.execute(
            "INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)",
            (username, generate_password_hash(password), role)
        )
        db.commit()
        flash(f"Usuario '{username}' creado correctamente.", "success")
    except sqlite3.IntegrityError:
        flash("Ya existe un usuario con ese nombre.", "error")

    return redirect(url_for("admin_users"))


@app.route("/admin/usuarios/<int:user_id>/toggle", methods=["POST"])
@admin_required
def admin_users_toggle(user_id):
    db = get_db()
    if user_id == session.get("user_id"):
        flash("No puedes desactivar tu propia cuenta.", "error")
        return redirect(url_for("admin_users"))

    user = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if user is None:
        flash("Usuario no encontrado.", "error")
        return redirect(url_for("admin_users"))

    nuevo_estado = 0 if user["activo"] else 1
    db.execute("UPDATE users SET activo = ? WHERE id = ?", (nuevo_estado, user_id))
    db.commit()
    flash("Usuario actualizado.", "success")
    return redirect(url_for("admin_users"))


# --------------------------------------------------------------------------
# Entrypoint
# --------------------------------------------------------------------------
init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
